# xrd.py
# Ferramenta para Análise e Simulação de Dados de Difração de Raios X (DRX)
# Versão 8.3 - Adicionada interface para Redução de Ruído com Transformada Wavelet
#
# CAMINHO REFATORADO: matfinder/tools/xrd/xrd.py
#

import sys
import os
import numpy as np
import json
import copy
import uuid
import shutil
import logging

# Importar PIL para redimensionamento de imagens (marca d'água)
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    logging.warning(ptr("PIL/Pillow não disponível. Marca d'água pode não funcionar corretamente."))
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QFileDialog, QMenu,
    QHeaderView, QMessageBox, QDialogButtonBox, QTabWidget, QWidget,
    QGroupBox, QFormLayout, QLineEdit, QLabel, QComboBox, QDoubleSpinBox,
    QSplitter, QListWidget, QListWidgetItem, QColorDialog, QCheckBox,
    QSpinBox, QAbstractItemView, QScrollArea, QSlider
)
from PySide6.QtCore import Qt, QPoint, Signal, Slot, QSize
from PySide6.QtGui import QAction, QPainter, QColor, QPen, QIcon

# --- Integração com Matplotlib para Gráficos ---
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.ticker import MultipleLocator, MaxNLocator
from matplotlib.widgets import RectangleSelector
import matplotlib.cm as cm
import matplotlib.image as mpimg
from matfinder.core.translator import ptr

# --- Importação para Exportar para Origin ---
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning(ptr("AVISO: Biblioteca 'openpyxl' não encontrada. A exportação para Origin (.xlsx) não funcionará."))


# --- Dependências para Cálculo de DRX ---
try:
    from pymatgen.core import Structure
    from pymatgen.analysis.diffraction.xrd import XRDCalculator

    PYMATGEN_AVAILABLE = True
except ImportError:
    PYMATGEN_AVAILABLE = False
    logging.warning(ptr("AVISO CRÍTICO: Biblioteca 'pymatgen' não encontrada. A simulação de DRX não funcionará."))

try:
    import pyobjcryst

    PYOBJCRYST_AVAILABLE = True
except ImportError:
    PYOBJCRYST_AVAILABLE = False
    logging.warning(ptr("AVISO: Biblioteca 'pyobjcryst' não encontrada. Usando 'pymatgen' como alternativa para simulação."))


# --- ALTERAÇÃO DE REATORAÇÃO: Importação de Módulos Locais ---
try:
    # 1. Importação Relativa (da mesma pasta 'xrd')
    from .xrd_math_tools import smooth_data, detect_peaks, denoise_with_wavelets, SCIPY_AVAILABLE, PYWAVELETS_AVAILABLE
    from .normalization_dialog import (NormalizationDialog, normalize_data, get_method_description,
                                       normalize_by_peak, NormalizeByPeakConfirmDialog)
    from .legend_dialog import LegendDialog
    from .reflection_dialog import ReflectionDialog
    from .structure_viewer import StructureViewer3D, Viewer3DToolsDialog, Viewer3DAnimationDialog

    # 2. Importações Absolutas (de outros pacotes 'matfinder')
    from matfinder.data.cif_editor_dialog import CifEditorDialog, MultiphaseEditorDialog
    from matfinder.ui_dialogs import BackgroundCorrectionDialog
except ImportError as e:
# --- FIM DA ALTERAÇÃO ---
    logging.warning(f"AVISO: Um módulo local não foi encontrado: {e}. Algumas funcionalidades podem não funcionar.")
    SCIPY_AVAILABLE = False
    PYWAVELETS_AVAILABLE = False


    class CifEditorDialog:
        pass


    class MultiphaseEditorDialog:
        pass


# Função auxiliar para configurar ícone do polvo em qualquer janela/dialog
def set_polvo_icon(window):
    """Configura o ícone do polvo em qualquer QWidget (QDialog, QMainWindow, etc)."""
    try:
        from PySide6.QtGui import QIcon
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = os.path.abspath(".")

        icon_path = os.path.join(base_path, "matfinder", "assets", "icons", "polvo.ico")

        if os.path.exists(icon_path):
            window.setWindowIcon(QIcon(icon_path))
        else:
            logging.debug(f"Ícone do polvo não encontrado em: {icon_path}")
    except Exception as e:
        logging.debug(f"Erro ao configurar ícone do polvo: {e}")


    class BackgroundCorrectionDialog:
        pass


    def smooth_data(*args):
        return None


    def detect_peaks(*args):
        return None, None


    def denoise_with_wavelets(*args):
        return None

# --- Classe customizada para QDoubleSpinBox com passo diferente na roda do mouse ---
class CustomDoubleSpinBox(QDoubleSpinBox):
    """QDoubleSpinBox customizado com passo diferente para roda do mouse."""

    def __init__(self, parent=None, wheel_step=None):
        super().__init__(parent)
        self._wheel_step = wheel_step  # Passo customizado para roda do mouse

    def wheelEvent(self, event):
        """Override do evento da roda do mouse para usar passo customizado."""
        if self._wheel_step is not None:
            # Obter direção da roda
            delta = event.angleDelta().y()
            if delta > 0:  # Roda para cima
                self.setValue(self.value() + self._wheel_step)
            elif delta < 0:  # Roda para baixo
                self.setValue(self.value() - self._wheel_step)
            event.accept()
        else:
            # Usar comportamento padrão se wheel_step não foi definido
            super().wheelEvent(event)


# --- Paleta de cores padrão para os gráficos ---
DEFAULT_PLOT_COLORS = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"
]


# --- FUNÇÕES DE PERFIL DE PICO ---
def gaussian(x, center, fwhm):
    sigma = fwhm / (2 * np.sqrt(2 * np.log(2)))
    profile = np.exp(-((x - center) ** 2) / (2 * sigma ** 2))
    # Normalizar para que o máximo seja 1
    return profile


def lorentzian(x, center, fwhm):
    gamma = fwhm / 2.0
    profile = gamma ** 2 / ((x - center) ** 2 + gamma ** 2)
    # Normalizar para que o máximo seja 1 (já está normalizado no centro x=center)
    return profile


def pseudo_voigt(x, center, fwhm, mixing=0.5):
    """mixing=0 -> Gaussiana pura, mixing=1 -> Lorentziana pura."""
    return (1 - mixing) * gaussian(x, center, fwhm) + mixing * lorentzian(x, center, fwhm)


class PlotCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)
        super().__init__(fig)
        self.setParent(parent)
        self.axes.set_xlabel("2θ (Graus)")
        self.axes.set_ylabel("Intensidade (Unid. arb.)")
        # NÃO usar tight_layout aqui - será aplicado apenas quando necessário
        # para evitar problemas com legendas customizadas



class PhaseDRXTool(QMainWindow):
    def __init__(self, parent=None, project_path=None):
        super().__init__(parent)

        self.current_project_path = project_path
        self.project_directory = os.path.dirname(project_path) if project_path else None
        self.is_dirty = False
        self.plot_items = []
        self.color_cycler_index = 0
        self.simulation_engine = "pymatgen" if PYMATGEN_AVAILABLE else "none"
        if PYOBJCRYST_AVAILABLE:
            self.simulation_engine = "pyobjcryst"

        self.history_stack = []
        self.history_index = -1

        self.group_settings = {"count": 1, "spacing": {}}

        self.zoom_selector = None
        self._original_x_lim = None
        self._original_y_lim = None
        self._is_zoomed = False

        # Modo de seleção de pico para normalização
        self._peak_selection_mode = False
        self._peak_selection_item_id = None
        self._peak_selection_cursor = None
        self.point_annotation = None
        self.annotation_dot = None
        self._is_dragging = False

        self.cif_editor_dialog_ref = None

        # Rastrear marca d'água (FigureImage object)
        self._watermark_image = None

        # Configurações e referência da legenda
        self.legend_settings = self._get_default_legend_settings()
        self._legend_ref = None  # Referência ao objeto Legend do matplotlib

        self.plot_settings = self._get_default_plot_settings()

        # Visualizador 3D: rastrear CIFs que já mostraram aviso
        self._warned_uncalculated_cifs = set()

        # Rastrear CIF pendente de renderização 3D
        self._pending_3d_cif_id = None
        self._has_pending_3d_changes = False

        # Flag para otimização: rastrear atualizações pendentes do matplotlib
        self._has_pending_plot_updates = False

        self._init_ui()
        self._connect_signals()

        if self.current_project_path:
            self.open_project(path=self.current_project_path)
        else:
            self._clear_session()

        if self.simulation_engine == "none":
            self.tab_widget.setTabEnabled(1, False)
            self.tab_widget.setTabToolTip(1, "As bibliotecas 'pymatgen' ou 'pyobjcryst' são necessárias.")

        self.set_dirty(False)

        # Configurar ícone da janela
        self._set_window_icon()

        # DECISÃO: Tela cheia SEMPRE, não redimensionável (evita erros de redimensionamento)
        # Remover botões de maximizar/minimizar, manter apenas fechar
        self.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)

        # Obter tamanho da tela e definir como fixo
        from PySide6.QtGui import QGuiApplication
        screen = QGuiApplication.primaryScreen().availableGeometry()
        self.setGeometry(screen)
        self.setFixedSize(screen.width(), screen.height())

        # Mostrar em tela cheia
        self.showMaximized()

    def resource_path(self, relative_path):
        """Retorna o caminho absoluto do recurso, funciona para dev e PyInstaller."""
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def _set_window_icon(self):
        """Configura o ícone da janela com o polvo."""
        icon_path = self.resource_path(os.path.join("matfinder", "assets", "icons", "polvo.ico"))
        if os.path.exists(icon_path):
            from PySide6.QtGui import QIcon
            self.setWindowIcon(QIcon(icon_path))
            logging.debug(f"Ícone do PhaseDRX carregado: {icon_path}")
        else:
            logging.warning(f"Ícone do PhaseDRX não encontrado em: {icon_path}")

    def _get_default_legend_settings(self):
        """Retorna configurações padrão da legenda."""
        return {
            "visible": True,
            "fontsize": 10,
            "fontweight": "normal",
            "fontstyle": "normal",
            "frameon": True,
            "reverse": False,
            "loc": "best",
            "draggable": True,
            "fancybox": True,
            "shadow": False,
            "framealpha": 0.9,
            "ncol": 1,
        }

    def _get_default_plot_settings(self):
        return {
            "x_lim": (0, 100), "y_lim": None, "x_ticks": 10.0, "y_ticks": None,
            "x_visible": True, "y_visible": True, "grid_visible": False,
            "x_label": "2θ (Graus)", "y_label": "Intensidade (Unid. arb.)",
            "x_label_fontsize": 12, "y_label_fontsize": 12,
            "x_label_bold": False, "y_label_bold": False,
            "x_label_italic": False, "y_label_italic": False,
            "global_offset": 0.0,
            "inter_group_offset": 0.2,
            "axes_linewidth": 1.3,
            "xtick_direction": 'in',
            "ytick_direction": 'in',
            "xtick_labelsize": 11,
            "ytick_labelsize": 11,
            "xtick_visible": True,
            "ytick_visible": False,  # Desativado por padrão
            "xtick_label_visible": True,
            "ytick_label_visible": False,  # Desativado por padrão
            "xtick_width": 1.0,
            "ytick_width": 1.0
        }

    def _init_ui(self):
        self.setWindowTitle(ptr("PhaseDRX"))
        self._create_menu_bar()
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        
        # Criar splitter principal e guardar referência
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self.main_splitter)
        
        # Estado do splitter para toggle
        self._splitter_collapsed = False
        self._splitter_original_sizes = [450, 750]

        control_widget = QWidget()
        control_layout = QVBoxLayout(control_widget)
        self.tab_widget = QTabWidget()
        control_layout.addWidget(self.tab_widget)

        self.exp_tab = QWidget()
        self._create_experimental_tab()
        self.tab_widget.addTab(self.exp_tab, ptr("Dados Experimentais"))

        self.sim_tab = QWidget()
        self._create_simulation_tab()
        self.tab_widget.addTab(self.sim_tab, ptr("Simulação Teórica (de CIF)"))

        self.main_splitter.addWidget(control_widget)

        # Desabilitar arrastar com o mouse no splitter
        self.main_splitter.setHandleWidth(0)
        self.main_splitter.setChildrenCollapsible(False)

        plot_panel = QWidget()
        plot_layout = QVBoxLayout(plot_panel)
        action_button_layout = QHBoxLayout()

        # Botão de toggle do splitter (expandir/recolher painel de controle)
        self.toggle_splitter_button = QPushButton("◀")
        self.toggle_splitter_button.setFixedSize(24, 24)
        self.toggle_splitter_button.setToolTip(ptr("Expandir/Recolher painel de controle"))
        self.toggle_splitter_button.clicked.connect(self._toggle_splitter)
        action_button_layout.addWidget(self.toggle_splitter_button)

        self.undo_button = QPushButton(ptr("◀ Desfazer"))
        self.redo_button = QPushButton(ptr("Refazer ▶"))
        action_button_layout.addWidget(self.undo_button)
        action_button_layout.addWidget(self.redo_button)
        action_button_layout.addStretch()

        self.edit_item_button = QPushButton("⚙️")
        self.edit_item_button.setToolTip(ptr("Editar parâmetros do item selecionado"))
        self.edit_item_button.setFixedSize(32, 32)
        self.edit_item_button.setEnabled(False)
        action_button_layout.addWidget(self.edit_item_button)

        # Botão de detectar picos removido da interface (funcionalidade mantida no backend)
        # self.peak_detect_button = QPushButton("Detectar Picos")
        # Botão Smooth movido para "Controles dos Dados Experimentais"
        self.save_plot_button = QPushButton(ptr("Salvar Gráfico..."))
        self.customize_plot_button = QPushButton(ptr("Personalizar Gráfico"))
        # action_button_layout.addWidget(self.peak_detect_button)
        action_button_layout.addWidget(self.save_plot_button)
        action_button_layout.addWidget(self.customize_plot_button)
        plot_layout.addLayout(action_button_layout)

        # Criar abas para gráfico 2D e visualizador 3D
        self.plot_tab_widget = QTabWidget()

        # Aba do gráfico 2D
        plot_2d_widget = QWidget()
        plot_2d_layout = QVBoxLayout(plot_2d_widget)
        plot_2d_layout.setContentsMargins(0, 0, 0, 0)
        self.plot_canvas = PlotCanvas(self, width=8, height=6)
        plot_2d_layout.addWidget(self.plot_canvas)
        self.plot_tab_widget.addTab(plot_2d_widget, ptr("Difratograma (2D)"))

        # Aba do visualizador 3D
        try:
            logging.info("Tentando criar visualizador 3D...")
            viewer_3d_widget = QWidget()
            viewer_3d_layout = QVBoxLayout(viewer_3d_widget)
            viewer_3d_layout.setContentsMargins(0, 0, 0, 0)
            viewer_3d_layout.setSpacing(0)

            self.structure_viewer = StructureViewer3D(self)
            viewer_3d_layout.addWidget(self.structure_viewer, stretch=1)

            # Adicionar botões de controle discretos na parte inferior
            render_button_layout = QHBoxLayout()
            render_button_layout.setContentsMargins(5, 3, 5, 3)
            render_button_layout.addStretch()

            # Botão de Ferramentas (configurações de visualização)
            self._tools_3d_button = QPushButton("⚙️")
            self._tools_3d_button.setFixedSize(32, 28)
            self._tools_3d_button.setToolTip(ptr("Configurações de visualização (tamanho dos átomos, célula unitária, etc.)"))
            self._tools_3d_button.clicked.connect(self._open_3d_tools_dialog)
            render_button_layout.addWidget(self._tools_3d_button)

            # Botão Renderizar
            self._render_3d_button = QPushButton(ptr("🔄 Renderizar"))
            self._render_3d_button.setMaximumWidth(120)
            self._render_3d_button.setEnabled(False)
            self._render_3d_button.setToolTip(ptr("Atualizar visualização 3D com modificações"))
            self._render_3d_button.clicked.connect(self._render_3d_structure)
            render_button_layout.addWidget(self._render_3d_button)

            # Botão de Animação
            self._animate_3d_button = QPushButton("▶️")
            self._animate_3d_button.setFixedSize(32, 28)
            self._animate_3d_button.setToolTip(ptr("Animações (rotação 360°, vibração)"))
            self._animate_3d_button.clicked.connect(self._open_animation_dialog)
            render_button_layout.addWidget(self._animate_3d_button)

            viewer_3d_layout.addLayout(render_button_layout)

            self.plot_tab_widget.addTab(viewer_3d_widget, ptr("Estrutura Cristalina (3D)"))
            logging.info("Visualizador 3D criado com sucesso!")
        except ImportError as e:
            logging.error(f"Erro de importação ao criar visualizador 3D: {e}")
            logging.error("Certifique-se de que 'pyqtgraph' e 'PyOpenGL' estão instalados.")
            self.structure_viewer = None
        except Exception as e:
            logging.error(f"Erro ao criar visualizador 3D: {e}", exc_info=True)
            self.structure_viewer = None

        plot_layout.addWidget(self.plot_tab_widget)

        # Conectar mudança de aba para renderização automática se houver pendências
        self.plot_tab_widget.currentChanged.connect(self._on_plot_tab_changed)

        # OTIMIZAÇÃO: Inicializar estado correto (Difratograma ativo, 3D inativo)
        if hasattr(self, 'structure_viewer') and self.structure_viewer is not None:
            if hasattr(self.structure_viewer, 'view_widget'):
                self.structure_viewer.view_widget.setEnabled(False)
            logging.debug("🚀 Inicialização: Difratograma ativo, Visualizador 3D pausado")

        self.main_splitter.addWidget(plot_panel)
        self.main_splitter.setSizes([450, 750])

        self._setup_zoom_selector()

    def _create_menu_bar(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("&Arquivo")
        new_action = QAction(ptr("&Novo Projeto..."), self, triggered=self.new_project)
        open_action = QAction(ptr("&Abrir Projeto..."), self, triggered=self.open_project)
        save_action = QAction(ptr("&Salvar Projeto"), self, triggered=self.save_project)
        save_as_action = QAction(ptr("Salvar Projeto &Como..."), self, triggered=lambda: self.save_project(save_as=True))
        
        # --- ALTERAÇÃO DE FUNCIONALIDADE: Adicionar Exportar para Origin ---
        export_origin_action = QAction(ptr("Exportar (.xlsx)..."), self, triggered=self.export_to_xlsx_for_origin)
        if not OPENPYXL_AVAILABLE:
            export_origin_action.setEnabled(False)
            export_origin_action.setToolTip(ptr("Instale 'openpyxl' para usar esta função"))
        # --- FIM DA ALTERAÇÃO ---

        exit_action = QAction(ptr("&Sair"), self, triggered=self.close)
        
        file_menu.addActions([new_action, open_action, save_action, save_as_action])
        file_menu.addSeparator()
        
        # --- ALTERAÇÃO DE FUNCIONALIDADE: Adicionar Ação ao Menu ---
        file_menu.addAction(export_origin_action) 
        file_menu.addSeparator()
        # --- FIM DA ALTERAÇÃO ---
        
        file_menu.addAction(exit_action)

    def _create_experimental_tab(self):
        layout = QVBoxLayout(self.exp_tab)
        load_group = QGroupBox(ptr("Controles dos Dados Experimentais"))
        load_layout = QVBoxLayout(load_group)
        self.load_button = QPushButton(ptr("Carregar Arquivo(s) Exp..."))
        self.normalize_button = QPushButton(ptr("Normalização..."))
        self.smooth_plot_button = QPushButton(ptr("Smooth"))

        load_controls_layout = QHBoxLayout()
        load_controls_layout.addWidget(self.load_button)
        load_controls_layout.addWidget(self.normalize_button)
        load_controls_layout.addWidget(self.smooth_plot_button)
        load_layout.addLayout(load_controls_layout)

        analysis_buttons_layout = QHBoxLayout()
        self.remove_background_button = QPushButton(ptr("Remover Fundo..."))

        # --- NOVO BOTÃO WAVELET ---
        self.denoise_wavelet_button = QPushButton(ptr("Redução de Ruído (Wavelet)"))
        if not PYWAVELETS_AVAILABLE:
            self.denoise_wavelet_button.setEnabled(False)
            self.denoise_wavelet_button.setToolTip(ptr("Instale a biblioteca 'PyWavelets' para usar esta função."))

        analysis_buttons_layout.addStretch()
        analysis_buttons_layout.addWidget(self.remove_background_button)
        analysis_buttons_layout.addWidget(self.denoise_wavelet_button)  # Adicionado à UI
        analysis_buttons_layout.addStretch()
        load_layout.addLayout(analysis_buttons_layout)

        layout.addWidget(load_group)
        exp_splitter = QSplitter(Qt.Orientation.Vertical)
        exp_list_group = QGroupBox(ptr("Arquivos Experimentais Carregados"))
        exp_list_layout = QVBoxLayout(exp_list_group)
        self.exp_list_widget = QListWidget()
        self.exp_list_widget.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.exp_list_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.exp_list_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        exp_list_layout.addWidget(self.exp_list_widget)
        exp_splitter.addWidget(exp_list_group)
        group_config_group = QGroupBox(ptr("Configuração de Grupos"))
        group_config_layout = QVBoxLayout(group_config_group)
        group_count_layout = QHBoxLayout()
        group_count_layout.addWidget(QLabel(ptr("Número de Grupos:")))
        self.num_groups_spin = QSpinBox()
        self.num_groups_spin.setRange(1, 20)
        self.num_groups_spin.setValue(1)
        group_count_layout.addWidget(self.num_groups_spin)
        group_count_layout.addStretch()
        group_config_layout.addLayout(group_count_layout)
        self.group_table_widget = QTableWidget()
        self.group_table_widget.setColumnCount(2)
        self.group_table_widget.setHorizontalHeaderLabels(["Grupo", "Nome do Gráfico"])
        self.group_table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.group_table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.group_table_widget.verticalHeader().setVisible(False)
        self.group_table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        group_config_layout.addWidget(self.group_table_widget)
        exp_splitter.addWidget(group_config_group)
        layout.addWidget(exp_splitter)
        exp_splitter.setSizes([200, 300])

    def _create_simulation_tab(self):
        layout = QHBoxLayout(self.sim_tab)
        cif_list_group = QGroupBox(ptr("Arquivos CIF Carregados"))
        cif_list_layout = QVBoxLayout(cif_list_group)
        self.cif_list_widget = QListWidget()
        self.cif_list_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.cif_list_widget.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.cif_list_widget.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        cif_list_layout.addWidget(self.cif_list_widget)
        layout.addWidget(cif_list_group, 1)
        params_group = QGroupBox(ptr("Parâmetros da Simulação"))
        params_layout = QVBoxLayout(params_group)
        form_layout = QFormLayout()
        self.cif_load_button = QPushButton(ptr("Carregar Arquivo(s) CIF..."))
        form_layout.addRow(self.cif_load_button)
        self.wavelength_combo = QComboBox()
        self.wavelength_combo.addItems(
            ["Cu Kα (1.54056 Å)", "Co Kα (1.78897 Å)", "Fe Kα (1.93604 Å)", "Mo Kα (0.70930 Å)", "Outro"])
        self.wavelength_custom_input = QLineEdit("1.54056")
        self.wavelength_custom_input.setVisible(False)
        wavelength_layout = QHBoxLayout()
        wavelength_layout.addWidget(self.wavelength_combo)
        wavelength_layout.addWidget(self.wavelength_custom_input)
        form_layout.addRow("Fonte de Radiação Padrão:", wavelength_layout)
        self.max_2theta_spin = QDoubleSpinBox()
        self.max_2theta_spin.setRange(10.0, 180.0)
        self.max_2theta_spin.setValue(100.0)
        self.max_2theta_spin.setSuffix(" °")
        form_layout.addRow("Máximo 2θ:", self.max_2theta_spin)
        self.peak_width_spin = QDoubleSpinBox()
        self.peak_width_spin.setRange(0.01, 1.0);
        self.peak_width_spin.setSingleStep(0.01)
        self.peak_width_spin.setDecimals(3);
        self.peak_width_spin.setValue(0.120)
        self.peak_width_spin.setSuffix(" °")
        self.peak_width_label = QLabel(ptr("Largura do Pico (FWHM):"))
        form_layout.addRow(self.peak_width_label, self.peak_width_spin)
        if self.simulation_engine != "pyobjcryst":
            self.peak_width_label.setVisible(False)
            self.peak_width_spin.setVisible(False)
        params_layout.addLayout(form_layout)
        params_layout.addStretch()

        action_buttons_layout = QHBoxLayout()
        self.combine_phases_button = QPushButton(ptr("Combinação Multifásica"))
        action_buttons_layout.addWidget(self.combine_phases_button)
        action_buttons_layout.addStretch()
        self.calculate_sim_button = QPushButton(ptr("Calcular Selecionado(s)"))
        action_buttons_layout.addWidget(self.calculate_sim_button)
        params_layout.addLayout(action_buttons_layout)

        layout.addWidget(params_group, 2)

    def _connect_signals(self):
        self.load_button.clicked.connect(self.load_experimental_files)
        self.normalize_button.clicked.connect(self.normalize_y_axis_selected)
        self.remove_background_button.clicked.connect(self.open_background_correction_dialog)
        # --- NOVA CONEXÃO DE SINAL ---
        self.denoise_wavelet_button.clicked.connect(self.open_wavelet_denoise_dialog)
        self.exp_list_widget.customContextMenuRequested.connect(self.show_exp_list_context_menu)
        self.exp_list_widget.model().rowsMoved.connect(self._on_plot_items_reordered)

        self.cif_load_button.clicked.connect(self.load_cif_files)
        self.cif_list_widget.customContextMenuRequested.connect(self.show_cif_list_context_menu)
        self.cif_list_widget.model().rowsMoved.connect(self._on_plot_items_reordered)
        self.cif_list_widget.itemSelectionChanged.connect(self._update_item_editor_button_state)
        self.cif_list_widget.itemSelectionChanged.connect(self._on_cif_selection_changed)

        self.combine_phases_button.clicked.connect(self.create_multiphase_combination)
        self.wavelength_combo.currentIndexChanged.connect(self.on_wavelength_change)
        self.calculate_sim_button.clicked.connect(self.run_simulation_for_selected)

        self.customize_plot_button.clicked.connect(self.open_plot_customization_dialog)
        self.save_plot_button.clicked.connect(self.open_save_plot_dialog)
        self.smooth_plot_button.clicked.connect(self.open_smooth_dialog)
        # Detectar picos removido da interface (funcionalidade mantida no backend)
        # self.peak_detect_button.clicked.connect(self.open_peak_detect_dialog)
        self.edit_item_button.clicked.connect(self.open_item_editor)

        self.num_groups_spin.valueChanged.connect(self.update_group_settings)
        self.group_table_widget.cellClicked.connect(self.on_group_cell_clicked)

        self.undo_button.clicked.connect(self.undo_action)
        self.redo_button.clicked.connect(self.redo_action)

        self.plot_canvas.mpl_connect('button_press_event', self._on_mouse_press)
        self.plot_canvas.mpl_connect('button_release_event', self._on_mouse_release)
        self.plot_canvas.mpl_connect('motion_notify_event', self._on_mouse_move)
        self.plot_canvas.mpl_connect('resize_event', self._on_canvas_resize)

    # --- NOVA FUNÇÃO PARA ABRIR O DIÁLOGO WAVELET ---
    def open_wavelet_denoise_dialog(self):
        """Abre o diálogo de redução de ruído por wavelet para o item selecionado."""
        selected_items = self.exp_list_widget.selectedItems()
        if len(selected_items) != 1:
            QMessageBox.warning(self, ptr("Seleção Inválida"),
                                ptr("Por favor, selecione um único arquivo experimental para a redução de ruído."))
            return

        list_item = selected_items[0]
        item_id = list_item.data(Qt.ItemDataRole.UserRole)
        plot_item = self._find_plot_item_by_id(item_id)

        if not plot_item or plot_item.get("data") is None:
            QMessageBox.critical(self, ptr("Erro de Dados"), ptr("O item selecionado não contém dados válidos."))
            return

        x_data, y_data = plot_item["data"]

        dialog = WaveletDenoiseDialog(x_data, y_data, self)
        dialog.denoising_applied.connect(
            lambda denoised_y: self.apply_wavelet_denoising(item_id, denoised_y)
        )
        dialog.exec()

    # --- NOVO SLOT PARA APLICAR O RESULTADO DO WAVELET ---
    @Slot(str, np.ndarray)
    def apply_wavelet_denoising(self, item_id, denoised_y_data):
        """Aplica os dados com ruído reduzido ao item do gráfico."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item:
            return

        self._add_state_to_history()

        plot_item["data"][1] = denoised_y_data

        # Adiciona um marcador para indicar que o tratamento foi feito
        if "analysis" not in plot_item:
            plot_item["analysis"] = {}
        plot_item["analysis"]["is_denoised"] = True

        # Picos antigos podem não ser mais válidos
        plot_item["analysis"]["peaks"] = None

        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def open_background_correction_dialog(self):
        """Abre o diálogo de remoção de fundo para o item experimental selecionado."""
        selected_items = self.exp_list_widget.selectedItems()
        if len(selected_items) != 1:
            QMessageBox.warning(self, ptr("Seleção Inválida"),
                                ptr("Por favor, selecione um único arquivo experimental para remover o fundo."))
            return

        list_item = selected_items[0]
        item_id = list_item.data(Qt.ItemDataRole.UserRole)
        plot_item = self._find_plot_item_by_id(item_id)

        if not plot_item or plot_item.get("data") is None:
            QMessageBox.critical(self, ptr("Erro de Dados"), ptr("O item selecionado não contém dados válidos."))
            return

        if plot_item.get("background_corrected", False):
            reply = QMessageBox.question(self, ptr("Refazer Correção"),
                                         "O fundo deste item já foi corrigido. Deseja refazer a operação?\n"
                                         "(Isto usará os dados originais novamente)",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                return
            x_data, _ = plot_item["data"]
            y_data = plot_item["original_y_data"]
        else:
            x_data, y_data = plot_item["data"]

        dialog = BackgroundCorrectionDialog(x_data, y_data, self)
        dialog.correction_applied.connect(
            lambda corrected_y: self.apply_background_correction(item_id, corrected_y)
        )
        dialog.exec()

    @Slot(str, np.ndarray)
    def apply_background_correction(self, item_id, corrected_y_data):
        """Aplica os dados de intensidade com fundo corrigido ao item do gráfico."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item:
            return

        self._add_state_to_history()

        if not plot_item.get("background_corrected", False):
            plot_item["original_y_data"] = plot_item["data"][1].copy()

        plot_item["data"][1] = corrected_y_data
        plot_item["background_corrected"] = True

        if "analysis" in plot_item:
            plot_item["analysis"]["peaks"] = None

        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def _update_item_editor_button_state(self):
        selected_items = self.cif_list_widget.selectedItems()
        self.edit_item_button.setEnabled(len(selected_items) == 1)

    def open_item_editor(self):
        selected_items = self.cif_list_widget.selectedItems()
        if len(selected_items) != 1:
            QMessageBox.warning(self, ptr("Seleção Inválida"), ptr("Por favor, selecione um único item para editar."))
            return

        item_id = selected_items[0].data(Qt.ItemDataRole.UserRole)
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item: return

        if plot_item.get("type") == "cif":
            self._open_cif_editor_dialog(plot_item)
        elif plot_item.get("type") == "multiphase":
            self._open_multiphase_editor_dialog(plot_item)

    def _open_cif_editor_dialog(self, plot_item):
        try:
            structure = Structure.from_str(plot_item["cif_content"], fmt="cif")
            calculator = XRDCalculator()
            pattern = calculator.get_pattern(structure, two_theta_range=(0, self.max_2theta_spin.value()))
            initial_peaks = pattern.x
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro ao Analisar CIF"),
                                 f"Não foi possível detectar os picos do CIF inicial:\n{e}")
            initial_peaks = []

        # Obter dados experimentais para auto-ajuste (primeiro experimental visível)
        experimental_data = None
        for item in self.plot_items:
            if item.get('type') == 'experimental' and item.get('visible', True):
                data = item.get('data')
                if data is not None and len(data) == 2:
                    experimental_data = (data[0], data[1])
                    break

        # Obter comprimento de onda atual
        try:
            wl_text = self.wavelength_combo.currentText()
            if wl_text == "Outro":
                wavelength = float(self.wavelength_custom_input.text())
            else:
                wavelength = float(wl_text.split('(')[1].split(' ')[0])
        except (ValueError, IndexError):
            wavelength = 1.5406

        max_2theta = self.max_2theta_spin.value()

        simulation_params = plot_item.get("simulation_params", {})
        self.cif_editor_dialog_ref = CifEditorDialog(
            plot_item["cif_content"], (initial_peaks, []), simulation_params, self,
            experimental_data=experimental_data, wavelength=wavelength,
            max_2theta=max_2theta
        )
        self.cif_editor_dialog_ref.simulationParametersModified.connect(
            lambda new_cif, params: self._handle_simulation_parameters_changed(plot_item["id"], new_cif, params)
        )
        # Conectar sinal para atualizar visualizador 3D com debounce
        self.cif_editor_dialog_ref.structureModified.connect(
            lambda new_cif: self._handle_structure_modified_for_3d(plot_item["id"], new_cif)
        )
        self.cif_editor_dialog_ref.exec()

    def _open_multiphase_editor_dialog(self, plot_item):
        component_info = []
        for i, comp_id in enumerate(plot_item.get("component_ids", [])):
            component = self._find_plot_item_by_id(comp_id)
            if component:
                info = {
                    "id": comp_id,
                    "label": component.get("label", "Desconhecido"),
                    "weight": plot_item.get("component_weights", [])[i]
                }
                component_info.append(info)

        dialog = MultiphaseEditorDialog(plot_item, component_info, self)
        dialog.parametersChanged.connect(self._handle_multiphase_parameters_changed)
        dialog.exec()

    def _handle_structure_modified_for_3d(self, item_id: str, new_cif_content: str):
        """Marca que há mudanças pendentes na estrutura CIF para renderização 3D."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item:
            return

        # Atualizar temporariamente o conteúdo do CIF
        plot_item["cif_content"] = new_cif_content

        # Marcar que há mudanças pendentes
        self._has_pending_3d_changes = True
        self._pending_3d_cif_id = item_id

        # Ativar botão de renderização se disponível
        if hasattr(self, '_render_3d_button') and self._render_3d_button:
            self._render_3d_button.setEnabled(True)
            self._render_3d_button.setStyleSheet("background-color: #FFA500; color: white; font-weight: bold;")
            logging.debug(f"Botão Renderizar ativado - mudanças pendentes para {plot_item.get('label')}")

    @Slot(str, str, dict)
    def _handle_simulation_parameters_changed(self, item_id: str, new_cif_content: str, sim_params: dict):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item: return

        self._add_state_to_history()
        plot_item["cif_content"] = new_cif_content
        plot_item["simulation_params"] = sim_params
        plot_item["is_modified"] = True
        self.run_simulation_for_item(plot_item)
        self.set_dirty()

    @Slot(str, list, dict)
    def _handle_multiphase_parameters_changed(self, item_id: str, new_weights: list, new_sim_params: dict):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("type") != "multiphase": return

        self._add_state_to_history()
        plot_item["component_weights"] = new_weights
        plot_item["simulation_params"] = new_sim_params
        self._recalculate_multiphase_item(plot_item)
        self.set_dirty()

    def create_multiphase_combination(self):
        selected_items = [self._find_plot_item_by_id(it.data(Qt.ItemDataRole.UserRole))
                          for it in self.cif_list_widget.selectedItems()]

        if len(selected_items) < 2:
            QMessageBox.warning(self, ptr("Seleção Insuficiente"),
                                ptr("Selecione pelo menos dois CIFs para criar uma combinação."))
            return

        for item in selected_items:
            if item.get("type") != "cif":
                QMessageBox.critical(self, ptr("Seleção Inválida"),
                                     f"O item '{item['label']}' não é um CIF e não pode ser combinado.")
                return
            if item.get("data") is None:
                QMessageBox.critical(self, ptr("Simulação Necessária"),
                                     f"O CIF '{item['label']}' ainda não foi simulado.\n"
                                     "Por favor, calcule todos os CIFs selecionados antes de combinar.")
                return

        self._add_state_to_history()

        num_items = len(selected_items)
        initial_weights = [1.0 / num_items] * num_items
        component_ids = [item['id'] for item in selected_items]
        labels = [item['label'].replace('.cif', '') for item in selected_items]
        new_label = "Comb: " + " + ".join(labels)

        new_item = {
            "id": str(uuid.uuid4()),
            "type": "multiphase",
            "label": new_label,
            "data": None,
            "visible": True,
            "color": self._get_next_color(),
            "style": "--", "linewidth": 2.0,
            "component_ids": component_ids,
            "component_weights": initial_weights,
            "simulation_params": {"profile": "pseudo-voigt", "fwhm": 0.120}
        }

        self._recalculate_multiphase_item(new_item)
        self.plot_items.append(new_item)

        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def _recalculate_multiphase_item(self, multiphase_item):
        component_ids = multiphase_item.get("component_ids", [])
        weights = multiphase_item.get("component_weights", [])
        sim_params = multiphase_item.get("simulation_params", {})

        if not component_ids or not weights: return

        try:
            wl_text = self.wavelength_combo.currentText()
            wavelength = float(self.wavelength_custom_input.text()) if wl_text == "Outro" else float(
                wl_text.split('(')[1].split(' ')[0])
            max_2theta = self.max_2theta_spin.value()
        except (ValueError, IndexError):
            return

        all_peaks = []
        for i, comp_id in enumerate(component_ids):
            component = self._find_plot_item_by_id(comp_id)
            if component and component.get("type") == "cif":
                try:
                    structure = Structure.from_str(component["cif_content"], fmt="cif")
                    calculator = XRDCalculator(wavelength=wavelength)
                    pattern = calculator.get_pattern(structure, two_theta_range=(0, max_2theta))

                    for peak_2theta, peak_intensity in zip(pattern.x, pattern.y):
                        all_peaks.append((peak_2theta, peak_intensity * weights[i]))
                except Exception as e:
                    logging.warning(f"Não foi possível obter picos para {component['label']}: {e}")
                    continue

        if not all_peaks:
            multiphase_item["data"] = None
            self._redraw_all_plots()
            return

        num_points = 4000
        ttheta_range = np.linspace(0, max_2theta, num_points)
        intensities = np.zeros(num_points)

        profile = sim_params.get("profile", "pseudo-voigt")
        fwhm = sim_params.get("fwhm", 0.120)

        for peak_2theta, peak_intensity in all_peaks:
            if profile == 'gaussiana':
                profile_func = gaussian(ttheta_range, peak_2theta, fwhm)
            elif profile == 'lorentziana':
                profile_func = lorentzian(ttheta_range, peak_2theta, fwhm)
            else:
                profile_func = pseudo_voigt(ttheta_range, peak_2theta, fwhm)
            intensities += peak_intensity * profile_func

        if intensities.max() > 0:
            intensities /= intensities.max()

        multiphase_item["data"] = [ttheta_range, intensities]
        self._redraw_all_plots()

    def _add_item_to_list(self, list_widget, item_data):
        list_item = QListWidgetItem()
        list_item.setData(Qt.ItemDataRole.UserRole, item_data["id"])
        list_widget.addItem(list_item)

        item_widget = QWidget()
        layout = QHBoxLayout(item_widget)
        layout.setContentsMargins(4, 2, 4, 2)

        visibility_button = QPushButton("👁️" if item_data.get("visible", True) else "⚪")
        visibility_button.setFlat(True)
        visibility_button.setFixedSize(24, 24)
        visibility_button.setToolTip(ptr("Ocultar/Mostrar no gráfico"))
        visibility_button.clicked.connect(lambda: self._toggle_item_visibility(item_data["id"]))

        label_text = item_data["label"]
        if item_data.get("is_modified"):
            label_text += " (M)"
        if item_data.get("type") == "multiphase":
            label_text += " (C)"
        if item_data.get("background_corrected"):
            label_text += " (B)"

        label_edit = QLineEdit(label_text)
        label_edit.setReadOnly(True)
        label_edit.setFrame(False)
        label_edit.setStyleSheet("background: transparent;")
        label_edit.setCursor(Qt.CursorShape.PointingHandCursor)

        # Desabilita o menu de contexto padrão do QLineEdit (Copy, Paste, Select All, etc.)
        label_edit.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)

        # Conecta clique único para selecionar o item
        label_edit.mousePressEvent = lambda event, lw=list_widget, it=list_item, le=label_edit: self._handle_label_click(event, lw, it, le)
        # Conecta duplo clique para ativar edição
        label_edit.mouseDoubleClickEvent = lambda event, le=label_edit, it=list_item: self._start_label_edit(le, it, event)
        label_edit.editingFinished.connect(lambda le=label_edit, it=list_item: self._finish_label_edit(le, it))

        layout.addWidget(visibility_button)
        layout.addWidget(label_edit)
        layout.addStretch()

        list_item.setSizeHint(item_widget.sizeHint())
        list_widget.setItemWidget(list_item, item_widget)

    def rename_plot_item(self, line_edit, list_item):
        new_text = line_edit.text().replace(" (M)", "").replace(" (C)", "").replace(" (B)", "").strip()
        item_id = list_item.data(Qt.ItemDataRole.UserRole)
        plot_item = self._find_plot_item_by_id(item_id)

        if plot_item and plot_item["label"] != new_text:
            self._add_state_to_history()
            plot_item["label"] = new_text
            self._repopulate_all_lists()
            self._repopulate_group_table()
            self._redraw_all_plots()
            self.set_dirty()

    def _handle_label_click(self, event, list_widget, list_item, line_edit):
        """Trata clique único no label para selecionar o item da lista"""
        # Se não estiver em modo de edição, seleciona o item
        if line_edit.isReadOnly():
            list_widget.setCurrentItem(list_item)
        # Chama o evento padrão para permitir outras interações
        QLineEdit.mousePressEvent(line_edit, event)

    def _start_label_edit(self, line_edit, list_item, event):
        """Ativa o modo de edição do nome ao dar duplo clique"""
        # Remove sufixos para edição limpa
        clean_text = line_edit.text().replace(" (M)", "").replace(" (C)", "").replace(" (B)", "")

        # Ativa edição
        line_edit.setReadOnly(False)
        line_edit.setFrame(True)
        line_edit.setStyleSheet("background: white; border: 1px solid #2196F3; padding: 2px;")
        line_edit.setCursor(Qt.CursorShape.IBeamCursor)
        line_edit.setText(clean_text)
        line_edit.setFocus()
        line_edit.selectAll()

        # Chama o evento padrão do QLineEdit
        QLineEdit.mouseDoubleClickEvent(line_edit, event)

    def _finish_label_edit(self, line_edit, list_item):
        """Finaliza a edição e salva o novo nome"""
        # Salva o novo nome
        self.rename_plot_item(line_edit, list_item)

        # Restaura o visual original
        line_edit.setReadOnly(True)
        line_edit.setFrame(False)
        line_edit.setStyleSheet("background: transparent;")
        line_edit.setCursor(Qt.CursorShape.PointingHandCursor)

    def _is_event_in_legend_zone(self, event, margin_pixels=20):
        """
        Verifica se o evento do mouse está na legenda ou em uma zona de segurança ao redor dela.

        Args:
            event: Evento do mouse do matplotlib
            margin_pixels: Margem em pixels ao redor da legenda (padrão: 20)

        Returns:
            bool: True se o evento está na zona da legenda, False caso contrário
        """
        if not self._legend_ref or not hasattr(event, 'x') or not hasattr(event, 'y'):
            return False

        try:
            # Obter bounding box da legenda em coordenadas de tela (pixels)
            bbox = self._legend_ref.get_window_extent(renderer=self.plot_canvas.figure.canvas.get_renderer())

            # Expandir bbox com margem de segurança
            x_min = bbox.x0 - margin_pixels
            x_max = bbox.x1 + margin_pixels
            y_min = bbox.y0 - margin_pixels
            y_max = bbox.y1 + margin_pixels

            # Verificar se o evento está dentro da zona expandida
            event_x = event.x
            event_y = event.y

            is_in_zone = (x_min <= event_x <= x_max) and (y_min <= event_y <= y_max)

            if is_in_zone:
                logging.debug(f"Evento na zona da legenda: ({event_x:.0f}, {event_y:.0f})")

            return is_in_zone

        except Exception as e:
            logging.debug(f"Erro ao verificar zona da legenda: {e}")
            return False

    def _setup_zoom_selector(self):
        """Configura o seletor retangular para zoom com visualização clara ESTILO WINDOWS XP."""
        # Variáveis para controle manual do retângulo de seleção
        self._zoom_rect = None  # Patch do retângulo visual
        self._zoom_start_point = None  # Ponto inicial do clique (x0, y0)
        self._last_zoom_point = None  # Última posição válida durante movimento
        self._is_selecting_zoom = False  # Flag de seleção ativa

        # Os eventos de mouse já estão conectados em _init_ui()
        # Vamos usar os handlers existentes: _on_mouse_press, _on_mouse_move, _on_mouse_release

        logging.debug("✅ Zoom selector manual configurado (estilo Windows XP)")

    def _toggle_splitter(self):
        """Alterna entre expandir e recolher o painel de controle."""
        if self._splitter_collapsed:
            # Expandir - voltar ao tamanho original
            self.main_splitter.setSizes(self._splitter_original_sizes)
            self.toggle_splitter_button.setText("◀")
            self.toggle_splitter_button.setToolTip(ptr("Recolher painel de controle"))
            self._splitter_collapsed = False
        else:
            # Recolher - salvar tamanho atual e colapsar
            self._splitter_original_sizes = self.main_splitter.sizes()
            # Colapsar painel de controle (primeiro widget) completamente
            total_width = sum(self._splitter_original_sizes)
            self.main_splitter.setSizes([0, total_width])
            self.toggle_splitter_button.setText("▶")
            self.toggle_splitter_button.setToolTip(ptr("Expandir painel de controle"))
            self._splitter_collapsed = True

    def _on_zoom_press(self, event):
        """Início da seleção de zoom - ao pressionar botão esquerdo."""
        # Ignorar se não for botão esquerdo ou não estiver no axes
        if event.button != 1 or not event.inaxes:
            return

        # Ignorar se estiver na zona da legenda
        if self._is_event_in_legend_zone(event, margin_pixels=30):
            logging.debug("✋ Seleção de zoom ignorada - clique na zona da legenda")
            return

        # Ignorar se estiver em modo de seleção de pico
        if self._peak_selection_mode:
            return

        # Iniciar seleção
        self._is_selecting_zoom = True
        self._zoom_start_point = (event.xdata, event.ydata)
        logging.debug(f"🔍 Início da seleção de zoom: ({event.xdata:.2f}, {event.ydata:.2f})")

    def _on_zoom_motion(self, event):
        """Durante a seleção - desenhar retângulo enquanto arrasta."""
        if not self._is_selecting_zoom or not event.inaxes:
            return

        if self._zoom_start_point is None:
            return

        x0, y0 = self._zoom_start_point
        x1, y1 = event.xdata, event.ydata

        if x1 is None or y1 is None:
            return

        # Remover retângulo anterior se existir
        if self._zoom_rect is not None:
            try:
                self._zoom_rect.remove()
            except:
                pass

        # Criar novo retângulo (estilo Windows XP)
        from matplotlib.patches import Rectangle
        width = x1 - x0
        height = y1 - y0

        self._zoom_rect = Rectangle(
            (x0, y0), width, height,
            linewidth=2,
            edgecolor='#0078D7',  # Azul Windows (moderno)
            facecolor='#CCE8FF',  # Azul claro Windows
            alpha=0.3,
            linestyle='-',
            fill=True
        )

        self.plot_canvas.axes.add_patch(self._zoom_rect)
        self.plot_canvas.draw_idle()

    def _on_zoom_release(self, event):
        """Fim da seleção - aplicar zoom."""
        if not self._is_selecting_zoom:
            return

        self._is_selecting_zoom = False

        # Remover retângulo visual
        if self._zoom_rect is not None:
            try:
                self._zoom_rect.remove()
            except:
                pass
            self._zoom_rect = None

        # Verificar se a seleção é válida
        if self._zoom_start_point is None or not event.inaxes:
            self._zoom_start_point = None
            self.plot_canvas.draw_idle()
            return

        x1, y1 = self._zoom_start_point
        x2, y2 = event.xdata, event.ydata

        if x2 is None or y2 is None:
            self._zoom_start_point = None
            self.plot_canvas.draw_idle()
            return

        # Verificar área mínima (10 pixels em cada direção)
        dx_pixels = abs(event.x - self.plot_canvas.axes.transData.transform((x1, y1))[0])
        dy_pixels = abs(event.y - self.plot_canvas.axes.transData.transform((x1, y1))[1])

        if dx_pixels < 10 or dy_pixels < 10:
            logging.debug("⚠️ Zoom cancelado - área muito pequena")
            self._zoom_start_point = None
            self.plot_canvas.draw_idle()
            return

        # Aplicar zoom
        if not self._is_zoomed:
            self._original_x_lim = self.plot_canvas.axes.get_xlim()
            self._original_y_lim = self.plot_canvas.axes.get_ylim()

        self.plot_canvas.axes.set_xlim(sorted((x1, x2)))
        self.plot_canvas.axes.set_ylim(sorted((y1, y2)))
        self._is_zoomed = True

        self._zoom_start_point = None
        self.plot_canvas.draw_idle()

        logging.debug(f"✅ Zoom aplicado: X[{min(x1,x2):.2f}, {max(x1,x2):.2f}], Y[{min(y1,y2):.2f}, {max(y1,y2):.2f}]")

    def _on_mouse_press(self, event):
        # PRIORIDADE: Modo de seleção de pico
        if self._peak_selection_mode and event.inaxes and event.button == 1:
            peak_x = event.xdata
            self._process_peak_selection(peak_x)
            return

        # Detectar duplo clique na legenda
        if event.dblclick and event.button == 1:
            if self._legend_ref and self._legend_ref.contains(event)[0]:
                self.open_legend_dialog()
                return

        # BOTÃO ESQUERDO: Iniciar seleção de zoom
        if event.inaxes and event.button == 1:
            self._is_dragging = False

            # Verificar se NÃO está na zona da legenda
            if not self._is_event_in_legend_zone(event, margin_pixels=30):
                # Iniciar seleção de zoom
                self._is_selecting_zoom = True
                self._zoom_start_point = (event.xdata, event.ydata)
                logging.debug(f"🔍 Início da seleção de zoom: ({event.xdata:.2f}, {event.ydata:.2f})")

        if event.inaxes and event.button == 3:
            if self.point_annotation:
                try:
                    self.point_annotation.set_visible(False)
                    if self.annotation_dot and len(self.annotation_dot) > 0:
                        self.annotation_dot[0].set_visible(False)
                except:
                    pass
                self.point_annotation = None
                self.annotation_dot = None

            if self._is_zoomed:
                self.plot_canvas.axes.set_xlim(self._original_x_lim)
                self.plot_canvas.axes.set_ylim(self._original_y_lim)
                self._is_zoomed = False
                self.plot_canvas.draw_idle()
                logging.debug("✅ Zoom desfeito com botão direito")

    def _on_mouse_move(self, event):
        # DESENHAR RETÂNGULO DE SELEÇÃO DE ZOOM (ESTILO WINDOWS XP)
        # Funciona mesmo quando o mouse está FORA do axes
        if event.button == 1:
            if event.inaxes:
                self._is_dragging = True

            if self._is_selecting_zoom and self._zoom_start_point is not None:
                x0, y0 = self._zoom_start_point

                # Se está dentro do axes, usar coordenadas normais
                if event.inaxes:
                    x1, y1 = event.xdata, event.ydata
                else:
                    # Mouse saiu do axes - clipar coordenadas aos limites
                    xlim = self.plot_canvas.axes.get_xlim()
                    ylim = self.plot_canvas.axes.get_ylim()

                    # Converter posição do mouse (pixels) para coordenadas de dados
                    try:
                        inv = self.plot_canvas.axes.transData.inverted()
                        x_data, y_data = inv.transform((event.x, event.y))

                        # Clipar aos limites do axes
                        x1 = max(xlim[0], min(xlim[1], x_data))
                        y1 = max(ylim[0], min(ylim[1], y_data))
                    except:
                        # Se falhar, manter última posição conhecida
                        if hasattr(self, '_last_zoom_point'):
                            x1, y1 = self._last_zoom_point
                        else:
                            return

                if x1 is not None and y1 is not None:
                    # Salvar última posição válida
                    self._last_zoom_point = (x1, y1)

                    # Remover retângulo anterior
                    if self._zoom_rect is not None:
                        try:
                            self._zoom_rect.remove()
                        except:
                            pass

                    # Criar novo retângulo (estilo Windows XP)
                    from matplotlib.patches import Rectangle
                    width = x1 - x0
                    height = y1 - y0

                    self._zoom_rect = Rectangle(
                        (x0, y0), width, height,
                        linewidth=1.0,        # Borda fina e minimalista
                        edgecolor='#0078D7',  # Azul Windows 10
                        facecolor='#CCE8FF',  # Azul claro Windows
                        alpha=0.15,           # Bem transparente (15%)
                        linestyle='-',
                        fill=True,
                        zorder=1000  # Desenhar por cima de tudo
                    )

                    self.plot_canvas.axes.add_patch(self._zoom_rect)
                    self.plot_canvas.draw_idle()

    def _on_mouse_release(self, event):
        # Salvar estado do dragging antes de resetar
        was_dragging = self._is_dragging

        # ========== CALCULAR MOVIMENTO (disponível em todo o escopo) ==========
        has_significant_movement = False
        if self._is_selecting_zoom and self._zoom_start_point is not None:
            try:
                x1, y1 = self._zoom_start_point

                # Obter posição final
                if event.inaxes:
                    x2, y2 = event.xdata, event.ydata
                elif hasattr(self, '_last_zoom_point'):
                    x2, y2 = self._last_zoom_point
                else:
                    x2, y2 = None, None

                # Calcular movimento em pixels
                if x2 is not None and y2 is not None:
                    pt1 = self.plot_canvas.axes.transData.transform((x1, y1))
                    pt2 = self.plot_canvas.axes.transData.transform((x2, y2))
                    dx_pixels = abs(pt2[0] - pt1[0])
                    dy_pixels = abs(pt2[1] - pt1[1])

                    # Movimento significativo = mais de 5 pixels em qualquer direção
                    has_significant_movement = (dx_pixels > 5 or dy_pixels > 5)
                    logging.debug(f"📏 Movimento: {dx_pixels:.1f}px H, {dy_pixels:.1f}px V → {'ZOOM' if has_significant_movement else 'CLIQUE'}")
            except Exception as e:
                logging.debug(f"Erro ao calcular movimento: {e}")

        # ========== PROCESSAR BOTÃO ESQUERDO ==========
        if event.button == 1:
            # CASO 1: ZOOM (movimento significativo)
            if self._is_selecting_zoom and self._zoom_start_point is not None and has_significant_movement:
                # Remover retângulo visual
                if self._zoom_rect is not None:
                    try:
                        self._zoom_rect.remove()
                    except:
                        pass
                    self._zoom_rect = None

                x1, y1 = self._zoom_start_point

                # Obter coordenadas finais
                if event.inaxes:
                    x2, y2 = event.xdata, event.ydata
                elif hasattr(self, '_last_zoom_point'):
                    x2, y2 = self._last_zoom_point
                    logging.debug(f"🎯 Mouse solto FORA do axes - usando última posição válida")
                else:
                    x2, y2 = None, None

                if x2 is not None and y2 is not None:
                    # Verificar área mínima (10 pixels)
                    try:
                        pt1 = self.plot_canvas.axes.transData.transform((x1, y1))
                        pt2 = self.plot_canvas.axes.transData.transform((x2, y2))
                        dx_pixels = abs(pt2[0] - pt1[0])
                        dy_pixels = abs(pt2[1] - pt1[1])

                        if dx_pixels >= 10 and dy_pixels >= 10:
                            # Aplicar zoom
                            if not self._is_zoomed:
                                self._original_x_lim = self.plot_canvas.axes.get_xlim()
                                self._original_y_lim = self.plot_canvas.axes.get_ylim()

                            self.plot_canvas.axes.set_xlim(sorted((x1, x2)))
                            self.plot_canvas.axes.set_ylim(sorted((y1, y2)))
                            self._is_zoomed = True

                            logging.debug(f"✅ Zoom aplicado: X[{min(x1,x2):.2f}, {max(x1,x2):.2f}], Y[{min(y1,y2):.2f}, {max(y1,y2):.2f}]")
                        else:
                            logging.debug("⚠️ Zoom cancelado - área muito pequena")
                    except:
                        pass

                # Limpar estado da seleção
                self._is_selecting_zoom = False
                self._zoom_start_point = None
                self._last_zoom_point = None
                self.plot_canvas.draw_idle()

            # CASO 2: CLIQUE ÚNICO (sem movimento significativo)
            elif event.inaxes:
                # Limpar retângulo se existir
                if self._zoom_rect is not None:
                    try:
                        self._zoom_rect.remove()
                    except:
                        pass
                    self._zoom_rect = None

                # Mostrar anotação
                self._toggle_annotation(event)
                logging.debug("👆 Clique único detectado - mostrando anotação")

                # Limpar estado de zoom
                self._is_selecting_zoom = False
                self._zoom_start_point = None
                self._last_zoom_point = None
                self.plot_canvas.draw_idle()

        self._is_dragging = False

    def _on_canvas_resize(self, event):
        """Reposiciona a logo quando a janela é redimensionada."""
        # Só reposiciona se não houver itens (logo deve estar visível)
        if not self.plot_items and self._watermark_image is not None:
            # Redesenhar tudo para reposicionar a logo
            self._redraw_all_plots()

    def _toggle_annotation(self, event):
        """Mostra/oculta anotação com informações do ponto clicado (2θ, Intensidade, d)."""
        if self.point_annotation:
            try:
                self.point_annotation.set_visible(False)
                if self.annotation_dot and len(self.annotation_dot) > 0:
                    for line in self.annotation_dot:
                        try:
                            line.set_visible(False)
                        except Exception:
                            pass
                self.plot_canvas.draw_idle()
            except Exception as e:
                logging.debug(f"Erro ao ocultar anotação: {e}")
            self.point_annotation = None
            self.annotation_dot = None
            return

        ax = self.plot_canvas.axes
        x, y = event.xdata, event.ydata
        if x is None or y is None:
            return

        # Calcular distância interplanar (d) usando Lei de Bragg: λ = 2d·sin(θ)
        # d = λ / (2·sin(θ))
        try:
            # Obter comprimento de onda atual
            wl_text = self.wavelength_combo.currentText()
            if wl_text == "Outro":
                wavelength = float(self.wavelength_custom_input.text())
            else:
                # Extrair wavelength do texto (formato: "Cu Kα (1.54056 Å)")
                wavelength = float(wl_text.split('(')[1].split(' ')[0])

            # Converter 2θ para θ (radianos)
            theta_rad = np.radians(x / 2.0)

            # Calcular d usando Lei de Bragg
            d_spacing = wavelength / (2.0 * np.sin(theta_rad))

            # Formatar texto da anotação
            annotation_text = f"2θ = {x:.2f}°\nInt = {y:.2f}\nd = {d_spacing:.4f} Å"
        except:
            # Se falhar, mostrar apenas 2θ e Intensidade
            annotation_text = f"2θ = {x:.2f}°\nInt = {y:.2f}"
            logging.debug("Não foi possível calcular distância interplanar")

        # Marcador no ponto (círculo vermelho minimalista)
        self.annotation_dot = ax.plot(x, y, 'o', color='#E74C3C', markersize=4, zorder=10)

        # Anotação moderna e minimalista
        self.point_annotation = ax.annotate(
            annotation_text,
            xy=(x, y),
            xytext=(12, 12),
            textcoords='offset points',
            fontsize=9,
            bbox=dict(
                boxstyle="round,pad=0.4",
                facecolor='white',
                edgecolor='#95A5A6',
                linewidth=0.8,
                alpha=0.95
            ),
            arrowprops=dict(
                arrowstyle='-',
                connectionstyle="arc3,rad=0",
                color='#95A5A6',
                linewidth=0.8
            ),
            ha='left',
            va='bottom'
        )
        self.plot_canvas.draw_idle()

    def set_dirty(self, dirty=True):
        self.is_dirty = dirty
        title = "PhaseDRX - "
        if self.current_project_path:
            title += os.path.basename(self.current_project_path)
        else:
            title += "Sessão Anônima"
        if self.is_dirty:
            title += "*"
        self.setWindowTitle(title)

    def new_project(self):
        """Cria um novo projeto, verificando antes se há alterações não salvas."""
        # Verifica se há alterações não salvas
        if not self.maybe_save():
            return

        # Importa o diálogo de projeto
        from matfinder.ui_dialogs import PhaseDRXProjectDialog

        # Mostra o diálogo para escolher tipo de projeto
        dialog = PhaseDRXProjectDialog(self)
        if not dialog.exec():
            return

        choice = dialog.choice
        project_path_to_open = None

        if choice == dialog.NEW_PROJECT:
            # Criar novo projeto
            project_name = dialog.project_name
            base_path = dialog.project_base_path
            project_dir = os.path.join(base_path, project_name)
            cif_dir = os.path.join(project_dir, "Cif")
            dados_dir = os.path.join(project_dir, "Dados")
            project_file = os.path.join(project_dir, f"{project_name}.mfpx")

            if os.path.exists(project_dir):
                reply = QMessageBox.question(
                    self, ptr("Pasta Existente"),
                    f"A pasta '{project_dir}' já existe.\n"
                    "Deseja abri-la como um projeto existente?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                if reply == QMessageBox.StandardButton.No:
                    return
                if not os.path.exists(project_file):
                    QMessageBox.warning(
                        self, ptr("Arquivo de Projeto Não Encontrado"),
                        f"A pasta existe, mas o arquivo de projeto '{project_file}' não foi encontrado."
                    )
                    return
                project_path_to_open = project_file
            else:
                try:
                    os.makedirs(cif_dir)
                    os.makedirs(dados_dir)
                    with open(project_file, 'w', encoding='utf-8') as f:
                        json.dump({}, f)
                    project_path_to_open = project_file
                    logging.info(f"Nova estrutura de projeto criada em: {project_dir}")
                except OSError as e:
                    QMessageBox.critical(
                        self, ptr("Erro ao Criar Projeto"),
                        f"Não foi possível criar a estrutura de pastas do projeto:\n{e}"
                    )
                    return

        elif choice == dialog.OPEN_PROJECT:
            # Abrir projeto existente
            path, _ = QFileDialog.getOpenFileName(
                self, "Abrir Projeto PhaseDRX", "",
                "Projetos MatFinder (*.mfpx);;Todos os Arquivos (*)"
            )
            if not path:
                return
            project_path_to_open = path

        elif choice == dialog.ANONYMOUS_SESSION:
            # Sessão anônima
            project_path_to_open = None

        # Limpa a sessão atual
        self._clear_session()

        # Configura o novo projeto
        self.current_project_path = project_path_to_open
        self.project_directory = os.path.dirname(project_path_to_open) if project_path_to_open else None

        # Se for um projeto existente, carrega os dados
        if project_path_to_open and os.path.exists(project_path_to_open):
            self.open_project(path=project_path_to_open)
        else:
            # Atualiza a interface
            self._restore_state_from_history()
            self.set_dirty(False)

            # Atualiza o título da janela
            title = "PhaseDRX - "
            if self.current_project_path:
                title += os.path.basename(self.current_project_path)
            else:
                title += "Sessão Anônima"
            if self.is_dirty:
                title += "*"
            self.setWindowTitle(title)

        logging.info(f"Novo projeto iniciado: {project_path_to_open or 'Sessão Anônima'}")

    def open_project(self, path=None):
        if not self.maybe_save(): return

        if not path:
            path, _ = QFileDialog.getOpenFileName(self, "Abrir Projeto", "",
                                                  "Projetos MatFinder (*.mfpx);;Todos os Arquivos (*)")
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    project_data = json.loads(content) if content else {}

                self.history_stack = []
                self.history_index = -1

                self.plot_items = project_data.get("plot_items", [])
                project_dir = os.path.dirname(path)

                for item in self.plot_items:
                    if item.get("data_is_embedded") and item.get("data"):
                        try:
                            item["data"][0] = np.array(item["data"][0])
                            item["data"][1] = np.array(item["data"][1])
                            if "original_y_data" in item and item["original_y_data"] is not None:
                                item["original_y_data"] = np.array(item["original_y_data"])
                        except (IndexError, TypeError) as e:
                            logging.warning(f"Aviso: Falha ao restaurar dados embutidos para '{item['label']}': {e}")
                            item["data"] = None

                    if item['type'] == 'cif':
                        modified_cif_rel_path = item.get('modified_cif_path')
                        original_cif_rel_path = item.get('path')

                        if modified_cif_rel_path:
                            full_modified_path = os.path.join(project_dir, modified_cif_rel_path)
                            if os.path.exists(full_modified_path):
                                with open(full_modified_path, 'r', encoding='utf-8') as cif_file:
                                    item['cif_content'] = cif_file.read()
                                item['is_modified'] = True
                            else:
                                logging.warning(
                                    f"Arquivo CIF modificado não encontrado em '{full_modified_path}'. Revertendo para o original.")
                                item['is_modified'] = False
                                item.pop('modified_cif_path', None)

                        if not item.get('is_modified') and original_cif_rel_path:
                            full_original_path = os.path.join(project_dir, original_cif_rel_path)
                            if os.path.exists(full_original_path):
                                with open(full_original_path, 'r', encoding='utf-8') as cif_file:
                                    item['cif_content'] = cif_file.read()
                            else:
                                logging.error(
                                    f"Arquivo CIF original não encontrado em '{full_original_path}' para o item '{item['label']}'.")

                        if original_cif_rel_path:
                            full_original_path = os.path.join(project_dir, original_cif_rel_path)
                            if os.path.exists(full_original_path):
                                with open(full_original_path, 'r', encoding='utf-8') as cif_file:
                                    item['original_cif_content'] = cif_file.read()

                for item in self.plot_items:
                    if item.get("type") == "multiphase":
                        self._recalculate_multiphase_item(item)

                self.plot_settings = self._get_default_plot_settings()
                project_plot_settings = project_data.get("plot_settings", {})
                self.plot_settings.update(project_plot_settings)

                self.group_settings = project_data.get("group_settings", {"count": 1, "spacing": {}})

                # Carregar configurações da legenda
                self.legend_settings = self._get_default_legend_settings()
                project_legend_settings = project_data.get("legend_settings", {})
                self.legend_settings.update(project_legend_settings)

                self._add_state_to_history()
                self._restore_state_from_history()

                self.current_project_path = path
                self.project_directory = os.path.dirname(path)
                self.set_dirty(False)
            except Exception as e:
                QMessageBox.critical(self, ptr("Erro ao Abrir"), f"Não foi possível carregar o arquivo de projeto:\n{e}")
                logging.exception("Erro crítico ao abrir projeto:")

    def save_project(self, save_as=False):
        path = self.current_project_path
        if save_as or not path:
            path, _ = QFileDialog.getSaveFileName(self, "Salvar Projeto Como...", "",
                                                  "Projetos MatFinder (*.mfpx);;Todos os Arquivos (*)")
        if not path: return False

        self.project_directory = os.path.dirname(path)
        cif_dir = os.path.join(self.project_directory, "Cif")
        modified_cif_dir = os.path.join(cif_dir, "modificados")
        dados_dir = os.path.join(self.project_directory, "Dados")
        os.makedirs(cif_dir, exist_ok=True)
        os.makedirs(modified_cif_dir, exist_ok=True)
        os.makedirs(dados_dir, exist_ok=True)

        try:
            items_to_save = copy.deepcopy(self.plot_items)

            def safe_filename(label, extension):
                base_name = os.path.splitext(label)[0]
                safe_name = "".join(c for c in base_name if c.isalnum() or c in (' ', '.', '_')).rstrip()
                return f"{safe_name}{extension}"

            for item in items_to_save:
                if item.get("data") is not None:
                    data_to_save = np.column_stack(item["data"])
                    if item['type'] == 'experimental':
                        exp_filename = safe_filename(item['label'], "_experimental.xy")
                        exp_target_path = os.path.join(dados_dir, exp_filename)
                        np.savetxt(exp_target_path, data_to_save, fmt='%.8f')
                    elif item['type'] in ['cif', 'multiphase']:
                        sim_filename = safe_filename(item['label'], ".dat")
                        sim_target_path = os.path.join(dados_dir, sim_filename)
                        np.savetxt(sim_target_path, data_to_save, fmt='%.8f')

                if item['type'] == 'cif':
                    cif_filename = safe_filename(item['label'], '.cif')

                    if item.get('original_cif_content'):
                        original_target_path = os.path.join(cif_dir, cif_filename)
                        with open(original_target_path, 'w', encoding='utf-8') as f:
                            f.write(item['original_cif_content'])
                        item['path'] = os.path.relpath(original_target_path, self.project_directory)

                    if item.get('is_modified') and item.get('cif_content'):
                        modified_target_path = os.path.join(modified_cif_dir, cif_filename)
                        with open(modified_target_path, 'w', encoding='utf-8') as f:
                            f.write(item['cif_content'])
                        item['modified_cif_path'] = os.path.relpath(modified_target_path, self.project_directory)
                    else:
                        item.pop('modified_cif_path', None)

                if item.get("data") is not None:
                    item["data"] = [item["data"][0].tolist(), item["data"][1].tolist()]
                    item["data_is_embedded"] = True
                    if "original_y_data" in item and item["original_y_data"] is not None:
                        item["original_y_data"] = item["original_y_data"].tolist()

                item.pop('cif_content', None)
                item.pop('original_cif_content', None)
                item.pop('original_path', None)

            settings_to_save = self.plot_settings.copy()
            if self._is_zoomed:
                settings_to_save['x_lim'] = self.plot_settings['x_lim']
                settings_to_save['y_lim'] = self.plot_settings['y_lim']

            project_data = {
                "version": "8.0",
                "plot_items": items_to_save,
                "plot_settings": settings_to_save,
                "group_settings": self.group_settings,
                "legend_settings": self.legend_settings
            }
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(project_data, f, indent=4)

            self.current_project_path = path
            self.set_dirty(False)
            QMessageBox.information(self, ptr("Projeto Salvo"), f"Projeto salvo com sucesso em:\n{path}")
            return True
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro ao Salvar"), f"Não foi possível salvar o arquivo de projeto:\n{e}")
            logging.exception("Erro crítico ao salvar projeto:")
            return False

    def maybe_save(self):
        if not self.is_dirty: return True
        ret = QMessageBox.warning(self, ptr("PhaseDRX"), ptr("Deseja salvar as suas alterações no projeto?"),
                                  QMessageBox.StandardButton.Save | QMessageBox.StandardButton.Discard | QMessageBox.StandardButton.Cancel)
        if ret == QMessageBox.StandardButton.Save:
            return self.save_project()
        elif ret == QMessageBox.StandardButton.Cancel:
            return False
        return True

    def _clear_session(self):
        self.plot_items = []
        self.color_cycler_index = 0
        self.group_settings = {"count": 1, "spacing": {}}
        self.plot_settings = self._get_default_plot_settings()
        self.history_stack = []
        self.history_index = -1
        self._add_state_to_history()
        self._restore_state_from_history()
        self.set_dirty(False)

    def closeEvent(self, event):
        # CORREÇÃO: Limpar visualizador 3D ANTES de mostrar diálogo para evitar erro OpenGL
        try:
            if hasattr(self, 'structure_viewer') and self.structure_viewer is not None:
                # Parar qualquer animação em execução
                if hasattr(self.structure_viewer, 'stop_animation'):
                    self.structure_viewer.stop_animation()

                # Parar timer do gizmo
                if hasattr(self.structure_viewer, '_gizmo_sync_timer'):
                    if self.structure_viewer._gizmo_sync_timer and self.structure_viewer._gizmo_sync_timer.isActive():
                        self.structure_viewer._gizmo_sync_timer.stop()

                # Desabilitar renderização OpenGL
                if hasattr(self.structure_viewer, 'view_widget'):
                    self.structure_viewer.view_widget.setEnabled(False)

                # Limpar estrutura para liberar recursos OpenGL
                if hasattr(self.structure_viewer, '_clear_structure'):
                    self.structure_viewer._clear_structure()

                logging.debug("✅ Visualizador 3D limpo antes de fechar")
        except Exception as e:
            logging.warning(f"Erro ao limpar visualizador 3D no closeEvent: {e}")

        # Agora sim verificar se deve salvar
        if self.maybe_save():
            event.accept()
        else:
            event.ignore()

    def resizeEvent(self, event):
        """Evento chamado quando a janela é redimensionada."""
        super().resizeEvent(event)
        # Redesenhar a legenda para que ela se ajuste às novas proporções
        if hasattr(self, '_legend_ref') and self._legend_ref:
            try:
                # A legenda com bbox_to_anchor usa coordenadas relativas (0-1)
                # então ela já se ajusta automaticamente ao redimensionamento
                self.plot_canvas.draw_idle()
            except:
                pass

    def load_experimental_files(self):
        file_filter = (
            "Arquivos de Difracao (*.xy *.dat *.raw *.csv *.xrdml *.brml *.ras *.asc *.int *.txt *.udf);;"
            "Bruker RAW (*.raw);;"
            "PANalytical XRDML (*.xrdml);;"
            "Bruker BRML (*.brml);;"
            "Rigaku RAS (*.ras);;"
            "Texto (*.xy *.dat *.asc *.int *.txt *.csv);;"
            "Todos os Arquivos (*)"
        )
        paths, _ = QFileDialog.getOpenFileNames(self, "Abrir Arquivo(s) de DRX", "", file_filter)
        if not paths: return

        self._add_state_to_history()
        for path in paths:
            try:
                ext = os.path.splitext(path)[1].lower()

                # Formatos binários ou especiais: usar raw_parser
                if ext in ('.raw', '.brml', '.xrdml', '.ras'):
                    from matfinder.tools.xrd.raw_parser import read_diffraction_file
                    x_data, y_data = read_diffraction_file(path)
                elif ext == '.csv':
                    from matfinder.tools.xrd.raw_parser import read_csv_file
                    x_data, y_data = read_csv_file(path)
                else:
                    # Formatos texto tradicionais (.xy, .dat, .asc, .int, .txt)
                    data_dict, headers = self._parse_data_file(path)
                    x_col_name, y_col_name = headers[0], headers[1]
                    x_data = np.array(data_dict[x_col_name], dtype=float)
                    y_data = np.array(data_dict[y_col_name], dtype=float)

                item_id = str(uuid.uuid4())
                # Remove extensão do nome do arquivo
                filename = os.path.basename(path)
                label_name = os.path.splitext(filename)[0]  # Remove .xy, .dat, etc.
                new_item = {
                    "id": item_id,
                    "type": "experimental",
                    "path": path,
                    "original_path": path,
                    "label": label_name,
                    "data": [x_data, y_data],
                    "color": self._get_next_color(),
                    "style": "-", "linewidth": 1.5, "visible": True,
                    "group_id": 0,
                    "analysis": {"peaks": None},
                    "is_smoothed": False,
                    "background_corrected": False
                }
                self.plot_items.append(new_item)
            except Exception as e:
                QMessageBox.critical(self, ptr("Erro ao Ler Arquivo"),
                                     f"Não foi possível processar o arquivo {os.path.basename(path)}:\n{e}")

        self._repopulate_all_lists()

        # Ajustar automaticamente o range do eixo X baseado nos dados experimentais carregados
        self._auto_adjust_x_range_for_experimental()

        # Ajustar automaticamente y_ticks baseado na escala dos dados (primeira vez apenas)
        if self.plot_settings.get("y_ticks") is None or self.plot_settings.get("y_ticks") == 0:
            self._auto_adjust_y_ticks_for_data()

        self._redraw_all_plots()
        self.set_dirty()

    def load_cif_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Abrir Arquivo(s) CIF", "", "CIF Files (*.cif)")
        if not paths: return

        self._add_state_to_history()
        for path in paths:
            if any(c.get("path") == path for c in self.plot_items): continue
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    cif_content = f.read()
                Structure.from_str(cif_content, fmt="cif")

                item_id = str(uuid.uuid4())
                # Remove extensão .cif do nome do arquivo
                filename = os.path.basename(path)
                label_name = os.path.splitext(filename)[0]  # Remove .cif
                new_item = {
                    "id": item_id, "type": "cif", "path": path,
                    "cif_content": cif_content,
                    "original_cif_content": cif_content,
                    "label": label_name,
                    "data": None,
                    "color": self._get_next_color(),
                    "style": "-", "linewidth": 1.0,
                    "visible": True, "group_id": 0,
                    "is_modified": False,
                    "simulation_params": {}
                }
                self.plot_items.append(new_item)
            except Exception as e:
                QMessageBox.critical(self, ptr("Erro ao Carregar CIF"),
                                     f"Não foi possível carregar ou processar o arquivo '{os.path.basename(path)}'.\n\nErro: {e}")

        self._repopulate_all_lists()
        self.set_dirty()

    def load_cif_from_data(self, cif_content: str, filename: str):
        if not PYMATGEN_AVAILABLE:
            QMessageBox.critical(self, ptr("Dependência Faltando"),
                                 ptr("A biblioteca 'pymatgen' é necessária para processar arquivos CIF."))
            return

        if any(c["label"] == filename for c in self.plot_items):
            QMessageBox.information(self, ptr("CIF Já Carregado"), f"O arquivo CIF '{filename}' já está na lista.")
            return

        self._add_state_to_history()
        try:
            Structure.from_str(cif_content, fmt="cif")
            item_id = str(uuid.uuid4())

            temp_dir = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else '.',
                                    'temp_cifs')
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, filename)
            with open(temp_path, 'w', encoding='utf-8') as f:
                f.write(cif_content)

            new_item = {
                "id": item_id, "type": "cif", "path": temp_path,
                "cif_content": cif_content,
                "original_cif_content": cif_content,
                "label": filename,
                "data": None, "color": self._get_next_color(), "style": "-",
                "linewidth": 1.0, "visible": True, "group_id": 0,
                "is_modified": False,
                "simulation_params": {}
            }
            self.plot_items.append(new_item)
            self._repopulate_all_lists()
            self.run_simulation_for_item(new_item)
            self.activateWindow();
            self.raise_()
            self.set_dirty()
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro no CIF"),
                                 f"O conteúdo do CIF recebido para '{filename}' é inválido.\n\nErro: {e}")

    def _parse_data_file(self, file_path):
        temp_data = []
        with open(file_path, 'r') as f:
            for line in f:
                parts = line.replace(',', ' ').strip().split()
                if not parts: continue
                try:
                    temp_data.append([float(p) for p in parts])
                except ValueError:
                    logging.warning(f"Ignorando linha não-numérica: {line.strip()}")
        if not temp_data: raise ValueError("Nenhum dado numérico válido foi encontrado.")
        num_cols = len(temp_data[0])
        headers = [chr(65 + i) for i in range(num_cols)]
        data_dict = {h: [row[i] for row in temp_data] for i, h in enumerate(headers)}
        return data_dict, headers

    def normalize_y_axis_selected(self):
        """Abre o diálogo de normalização e aplica o método selecionado aos dados experimentais."""
        selected_items = self.exp_list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, ptr("Nenhuma Seleção"),
                              ptr("Selecione um ou mais arquivos experimentais da lista para normalizar."))
            return

        # Abrir diálogo de seleção do método
        dialog = NormalizationDialog(self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return  # Usuário cancelou

        method = dialog.get_selected_method()
        method_name = get_method_description(method)

        self._add_state_to_history()

        # Aplicar normalização aos itens selecionados
        success_count = 0
        error_count = 0

        for list_item in selected_items:
            item_id = list_item.data(Qt.ItemDataRole.UserRole)
            plot_item = self._find_plot_item_by_id(item_id)
            if not plot_item or plot_item["data"] is None:
                continue

            try:
                x_data, y_data = plot_item["data"]
                normalized_y = normalize_data(y_data, method=method)
                plot_item["data"][1] = normalized_y
                success_count += 1
            except Exception as e:
                error_count += 1
                QMessageBox.critical(self, ptr("Erro de Normalização"),
                                   f"Não foi possível normalizar '{plot_item['label']}':\n{e}")

        # Redesenhar gráfico e atualizar estado
        if success_count > 0:
            self._redraw_all_plots()
            self.set_dirty()

            # Mostrar mensagem de sucesso
            if error_count == 0:
                QMessageBox.information(self, ptr("Normalização Concluída"),
                                      f"{success_count} arquivo(s) normalizado(s) com sucesso usando:\n{method_name}")
            else:
                QMessageBox.warning(self, ptr("Normalização Parcial"),
                                  f"{success_count} arquivo(s) normalizado(s), {error_count} com erro.")
        self.set_dirty()

    def run_simulation_for_selected(self):
        selected_list_items = self.cif_list_widget.selectedItems()
        if not selected_list_items:
            QMessageBox.warning(self, ptr("Nenhuma Seleção"), ptr("Selecione um ou mais arquivos CIF da lista para calcular."))
            return

        self._add_state_to_history()
        for list_item in selected_list_items:
            item_id = list_item.data(Qt.ItemDataRole.UserRole)
            plot_item = self._find_plot_item_by_id(item_id)
            if plot_item and plot_item.get("type") == "cif":
                self.run_simulation_for_item(plot_item)

    def run_simulation_for_item(self, plot_item):
        if self.simulation_engine == "none": return
        try:
            wl_text = self.wavelength_combo.currentText()
            wavelength = float(self.wavelength_custom_input.text()) if wl_text == "Outro" else float(
                wl_text.split('(')[1].split(' ')[0])
            max_2theta, peak_width = self.max_2theta_spin.value(), self.peak_width_spin.value()
        except (ValueError, IndexError) as e:
            QMessageBox.critical(self, ptr("Erro de Parâmetro"), f"Valor inválido: {e}")
            return

        try:
            if self.simulation_engine == "pyobjcryst":
                ttheta, intensity = self._calculate_with_pyobjcryst(plot_item, wavelength, max_2theta, peak_width)
            else:
                ttheta, intensity = self._calculate_with_pymatgen(plot_item, wavelength, max_2theta)
            plot_item["data"] = [ttheta, intensity]

            self._redraw_all_plots()
            self.set_dirty()

            # IMPORTANTE: NÃO renderizar automaticamente a estrutura 3D após calcular
            # O usuário deve clicar explicitamente em "Visualizar Estrutura"
            logging.debug(f"CIF calculado: {plot_item['label']} - estrutura 3D NÃO renderizada automaticamente")
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro na Simulação"),
                                 f"Ocorreu um erro ao calcular para {plot_item['label']}:\n{e}")

    def _calculate_with_pymatgen(self, cif_info, wavelength, max_2theta_deg):
        structure = Structure.from_str(cif_info["cif_content"], fmt="cif")
        calculator = XRDCalculator(wavelength=wavelength)
        pattern = calculator.get_pattern(structure, two_theta_range=(0, max_2theta_deg))

        num_points = 4000
        ttheta_range = np.linspace(0, max_2theta_deg, num_points)
        intensities = np.zeros(num_points)

        sim_params = cif_info.get("simulation_params", {})
        profile = sim_params.get("profile", "pseudo-voigt")
        fwhm = sim_params.get("fwhm", 0.120)
        apply_mode = sim_params.get("apply_mode", "Padrão Inteiro")
        ranges = sim_params.get("ranges", [])

        for peak_2theta, peak_intensity in zip(pattern.x, pattern.y):
            apply_convolution = False
            if apply_mode == "Padrão Inteiro":
                apply_convolution = True
            else:
                for r_min, r_max in ranges:
                    if r_min <= peak_2theta <= r_max:
                        apply_convolution = True
                        break

            if apply_convolution:
                if profile == 'gaussiana':
                    profile_func = gaussian(ttheta_range, peak_2theta, fwhm)
                elif profile == 'lorentziana':
                    profile_func = lorentzian(ttheta_range, peak_2theta, fwhm)
                else:
                    profile_func = pseudo_voigt(ttheta_range, peak_2theta, fwhm)
                intensities += peak_intensity * profile_func
            else:
                idx = np.argmin(np.abs(ttheta_range - peak_2theta))
                intensities[idx] += peak_intensity

        # Normalizar para [0, 1] para manter mesma escala que experimental normalizado
        max_intensity = intensities.max()
        if max_intensity > 0:
            intensities /= max_intensity

        return ttheta_range, intensities

    def _calculate_with_pyobjcryst(self, cif_info, wavelength, max_2theta_deg, peak_width_deg):
        temp_cif_path = "temp_cif_for_pyobjcryst.cif"
        with open(temp_cif_path, "w", encoding='utf-8') as f:
            f.write(cif_info["cif_content"])
        crystal = pyobjcryst.loadCrystal(temp_cif_path)
        os.remove(temp_cif_path)
        pattern_data = pyobjcryst.powderpattern.PowderPattern()
        pattern_data.SetWavelength(wavelength)
        numpoints, deg_to_rad = 4000, np.pi / 180.0
        ttheta_rad = np.linspace(0, max_2theta_deg * deg_to_rad, numpoints)
        pattern_data.SetPowderPatternX(ttheta_rad)
        diff_data = pattern_data.AddPowderPatternDiffraction(crystal)
        diff_data.SetReflectionProfilePar(pyobjcryst.powderpattern.ReflectionProfileType.PROFILE_PSEUDO_VOIGT,
                                          (peak_width_deg * deg_to_rad) ** 2)
        intensity = pattern_data.GetPowderPatternCalc()
        if intensity.max() > 0: intensity /= intensity.max()
        return ttheta_rad / deg_to_rad, intensity

    def _repopulate_all_lists(self):
        self.exp_list_widget.clear()
        self.cif_list_widget.clear()

        exp_items = [p for p in self.plot_items if p['type'] == 'experimental']
        cif_items = [p for p in self.plot_items if p['type'] in ['cif', 'multiphase']]

        for item in exp_items: self._add_item_to_list(self.exp_list_widget, item)
        for item in cif_items: self._add_item_to_list(self.cif_list_widget, item)

        self._repopulate_group_table()

    def _toggle_item_visibility(self, item_id):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item: return

        self._add_state_to_history()
        plot_item["visible"] = not plot_item.get("visible", True)
        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def _repopulate_group_table(self):
        self.group_table_widget.setRowCount(0)
        self.group_table_widget.setRowCount(len(self.plot_items))

        for i, item in enumerate(self.plot_items):
            group_id = item.get("group_id", 0)

            group_cell_widget = QWidget()
            layout = QHBoxLayout(group_cell_widget)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            circle_label = QLabel()
            color = DEFAULT_PLOT_COLORS[group_id % len(DEFAULT_PLOT_COLORS)]
            circle_label.setStyleSheet(f"background-color: {color}; border-radius: 8px; border: 1px solid black;")
            circle_label.setFixedSize(16, 16)
            layout.addWidget(circle_label)
            self.group_table_widget.setCellWidget(i, 0, group_cell_widget)

            name_item = QTableWidgetItem(item["label"])
            name_item.setData(Qt.ItemDataRole.UserRole, item["id"])
            self.group_table_widget.setItem(i, 1, name_item)

        self.group_table_widget.resizeRowsToContents()

    def _on_plot_items_reordered(self):
        self._add_state_to_history()
        exp_ids = [self.exp_list_widget.item(i).data(Qt.ItemDataRole.UserRole) for i in
                   range(self.exp_list_widget.count())]
        cif_ids = [self.cif_list_widget.item(i).data(Qt.ItemDataRole.UserRole) for i in
                   range(self.cif_list_widget.count())]

        all_ids_in_order = exp_ids + cif_ids

        item_map = {item['id']: item for item in self.plot_items}

        self.plot_items = [item_map[id] for id in all_ids_in_order if id in item_map]

        self._repopulate_group_table()
        self._redraw_all_plots()
        self.set_dirty()

    def update_group_settings(self):
        self._add_state_to_history()
        self.group_settings["count"] = self.num_groups_spin.value()
        for item in self.plot_items:
            if item.get("group_id", 0) >= self.group_settings["count"]:
                item["group_id"] = 0
        self._repopulate_group_table()
        self._redraw_all_plots()
        self.set_dirty()

    def on_group_cell_clicked(self, row, column):
        if column == 0:
            item_id = self.group_table_widget.item(row, 1).data(Qt.ItemDataRole.UserRole)
            plot_item = self._find_plot_item_by_id(item_id)
            if plot_item:
                self._add_state_to_history()
                current_group = plot_item.get("group_id", 0)
                num_groups = self.group_settings.get("count", 1)
                plot_item["group_id"] = (current_group + 1) % num_groups
                self._repopulate_group_table()
                self._redraw_all_plots()
                self.set_dirty()

    def _auto_adjust_x_range_for_experimental(self):
        """Ajusta automaticamente o range do eixo X baseado nos dados experimentais."""
        experimental_items = [item for item in self.plot_items
                             if item.get("type") == "experimental"
                             and item.get("data") is not None]

        if not experimental_items:
            return

        min_2theta = float('inf')
        max_2theta = float('-inf')

        for item in experimental_items:
            x_data = item["data"][0]
            if len(x_data) > 0:
                min_2theta = min(min_2theta, np.min(x_data))
                max_2theta = max(max_2theta, np.max(x_data))

        if min_2theta != float('inf') and max_2theta != float('-inf'):
            # Usar apenas a parte inteira
            min_int = int(np.floor(min_2theta))
            max_int = int(np.ceil(max_2theta))

            # Ajustar o x_lim no plot_settings
            self.plot_settings["x_lim"] = (min_int, max_int)

    def _auto_adjust_y_ticks_for_data(self):
        """
        Ajusta automaticamente y_ticks baseado na escala dos dados carregados.
        Calcula um valor adequado para evitar geração excessiva de ticks.
        """
        # Coletar todos os dados Y de todos os itens visíveis
        all_y_values = []

        for item in self.plot_items:
            if item.get("visible", True) and item.get("data") is not None:
                y_data = item["data"][1]
                if len(y_data) > 0:
                    all_y_values.extend(y_data)

        if not all_y_values:
            return

        # Calcular a escala dos dados
        y_max = np.max(all_y_values)
        y_min = np.min(all_y_values)
        y_range = y_max - y_min

        # Calcular um y_ticks adequado baseado na ordem de magnitude
        # Objetivo: ter aproximadamente 5-10 ticks no eixo Y
        # CRÍTICO: Prevenir geração de milhares de ticks que trava o matplotlib
        if y_range > 0:
            # Proteção contra range muito pequeno
            if y_range < 1e-6:
                self.plot_settings["y_ticks"] = 0.1
                logging.warning(f"Y-range muito pequeno ({y_range:.2e}), usando y_ticks padrão: 0.1")
                return

            # Calcular ordem de magnitude
            magnitude = 10 ** np.floor(np.log10(y_range))

            # Tentar valores comuns: 1, 2, 5, 10, 20, 50, 100, 200, 500, 1000...
            nice_steps = [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000, 2000, 5000, 10000]
            nice_steps_scaled = [step * magnitude / 100 for step in nice_steps]

            # Encontrar o step que resulta em ~5-10 ticks
            target_ticks = 8
            best_step = magnitude

            for step in nice_steps_scaled:
                if step > 0:
                    num_ticks = y_range / step
                    if 5 <= num_ticks <= 15:  # Faixa aceitável de ticks
                        best_step = step
                        break

            # Garantir que o step seja >= 0.1 para evitar problemas
            if best_step < 0.1:
                best_step = max(0.1, y_range / 10)

            # Arredondar para um valor "bonito"
            if best_step >= 1:
                best_step = round(best_step)
            elif best_step >= 0.1:
                best_step = round(best_step, 1)
            else:
                best_step = round(best_step, 2)

            # VALIDAÇÃO FINAL CRÍTICA: Garantir que não gerará mais de 1000 ticks
            # (limite MAXTICKS do matplotlib é 1000)
            estimated_ticks = y_range / best_step
            if estimated_ticks > 500:  # Usar margem de segurança (50% do limite)
                # Recalcular para ter no máximo 10 ticks
                best_step = y_range / 10
                logging.warning(
                    f"Y-ticks ajustado para prevenir excesso: {best_step:.2f} "
                    f"(teria gerado {estimated_ticks:.0f} ticks)"
                )

            self.plot_settings["y_ticks"] = best_step

            logging.info(f"Y-ticks ajustado automaticamente: {best_step} (range dos dados: {y_min:.1f} a {y_max:.1f})")

    def _draw_watermark(self, ax):
        """
        Desenha a marca d'água PhaseDRX.png centralizada no gráfico quando não há dados.
        A imagem permanece centralizada VISUALMENTE no canvas, independente dos limites dos eixos.
        Tamanho fixo e centralização usando coordenadas da figura (figure coordinates).
        """
        try:
            # Caminho da logo (CORREÇÃO: no app congelado os assets ficam sob
            # 'matfinder/assets/...' dentro do _MEIPASS; faltava o prefixo
            # 'matfinder', então a marca d'água nunca aparecia no .exe/instalador).
            if getattr(sys, 'frozen', False):
                # Executável compilado: _MEIPASS/matfinder/assets/logos/...
                logo_path = os.path.join(sys._MEIPASS, 'matfinder', 'assets', 'logos', 'PhaseDRX.png')
            else:
                # Desenvolvimento: base_path == .../matfinder
                base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                logo_path = os.path.join(base_path, 'assets', 'logos', 'PhaseDRX.png')

            if os.path.exists(logo_path):
                # Carregar a imagem
                img = mpimg.imread(logo_path)

                # Usar figimage ao invés de imshow para coordenadas absolutas da figura
                # Isso garante que a marca d'água fique sempre centralizada visualmente
                # independente dos limites dos eixos (xlim/ylim)

                fig = ax.figure

                # Obter tamanho do canvas em pixels
                bbox = ax.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
                width_inches = bbox.width
                height_inches = bbox.height

                # Converter para pixels
                dpi = fig.dpi
                canvas_width_px = width_inches * dpi
                canvas_height_px = height_inches * dpi

                # Dimensões da imagem original
                img_height, img_width = img.shape[:2]
                aspect_ratio = img_width / img_height

                # Definir tamanho fixo: 40% da largura do canvas ou 40% da altura (o que for menor)
                max_width = canvas_width_px * 0.4
                max_height = canvas_height_px * 0.4

                # Calcular tamanho mantendo proporção
                if max_width / aspect_ratio <= max_height:
                    # Limita pela largura
                    new_width = max_width
                    new_height = max_width / aspect_ratio
                else:
                    # Limita pela altura
                    new_height = max_height
                    new_width = max_height * aspect_ratio

                # Redimensionar imagem se PIL disponível
                if PIL_AVAILABLE:
                    pil_img = Image.fromarray((img * 255).astype('uint8'))
                    pil_img = pil_img.resize((int(new_width), int(new_height)), Image.Resampling.LANCZOS)
                    img_resized = np.array(pil_img) / 255.0
                else:
                    # Se PIL não disponível, usar imagem original (pode ficar em tamanho incorreto)
                    img_resized = img

                # Calcular posição central (em coordenadas da figura, origem no canto inferior esquerdo)
                # Obter posição do axes na figura
                ax_bbox = ax.get_position()
                fig_width, fig_height = fig.get_size_inches()

                # Posição do axes em pixels
                ax_left_px = ax_bbox.x0 * fig_width * dpi
                ax_bottom_px = ax_bbox.y0 * fig_height * dpi
                ax_width_px = ax_bbox.width * fig_width * dpi
                ax_height_px = ax_bbox.height * fig_height * dpi

                # Centro do axes em pixels
                center_x_px = ax_left_px + ax_width_px / 2
                center_y_px = ax_bottom_px + ax_height_px / 2

                # Posição da imagem (canto inferior esquerdo)
                img_x = center_x_px - new_width / 2
                img_y = center_y_px - new_height / 2

                # Adicionar imagem com transparência usando figimage
                # figimage usa coordenadas absolutas em pixels da figura
                # IMPORTANTE: Armazenar referência para poder remover depois
                self._watermark_image = fig.figimage(img_resized, xo=img_x, yo=img_y, alpha=0.15, zorder=0, origin='upper')
                logging.debug(f"Marca d'água adicionada: {self._watermark_image}")

            else:
                logging.warning(f"Logo PhaseDRX.png não encontrada em: {logo_path}")

        except Exception as e:
            logging.error(f"Erro ao desenhar marca d'água: {e}")

    def _redraw_all_plots(self):
        ax = self.plot_canvas.axes

        current_xlim = None
        current_ylim = None
        if self._is_zoomed:
            current_xlim = ax.get_xlim()
            current_ylim = ax.get_ylim()

        ax.clear()

        # IMPORTANTE: Remover marca d'água existente ANTES de verificar se deve adicionar nova
        # Usar referência armazenada para garantir remoção correta
        if self._watermark_image is not None:
            try:
                self._watermark_image.remove()
                logging.debug(f"Marca d'água removida: {self._watermark_image}")
            except:
                pass  # Já foi removida ou não existe mais
            self._watermark_image = None

        # Verificar se há QUALQUER item nas listas (visível ou não)
        # Marca d'água só aparece se as listas estiverem completamente vazias
        if not self.plot_items:  # Nenhum item em nenhuma lista (vazia)
            self._draw_watermark(ax)
        else:
            logging.debug(f"Marca d'água NÃO adicionada - há {len(self.plot_items)} item(s) nas listas")

        # Agrupar itens visíveis para plotagem
        groups = {}
        for item in self.plot_items:
            if not item.get("visible", True) or item.get("data") is None:
                continue
            group_id = item.get("group_id", 0)
            if group_id not in groups:
                groups[group_id] = []
            groups[group_id].append(item)


        global_offset = self.plot_settings.get("global_offset", 0.0)
        inter_group_offset = self.plot_settings.get("inter_group_offset", 0.2)

        current_y_offset = global_offset
        sorted_group_ids = sorted(groups.keys())

        for i, group_id in enumerate(sorted_group_ids):
            items_in_group = groups[group_id]
            intra_group_offset = self.group_settings.get("spacing", {}).get(str(group_id), 0.0)

            if i > 0:
                current_y_offset += inter_group_offset

            for j, item in enumerate(items_in_group):
                x_data, y_data = item["data"]


                offset = current_y_offset + (j * intra_group_offset)

                ax.plot(x_data, y_data + offset,
                        label=item.get("label", ""),
                        linestyle=item.get("style", "-"),
                        color=item.get("color", "k"),
                        linewidth=item.get("linewidth", 1.0))

                if item.get("type") == "experimental" and item.get("analysis", {}).get("peaks"):
                    peak_x, peak_y = item["analysis"]["peaks"]
                    peak_y_array = np.array(peak_y)
                    ax.plot(peak_x, peak_y_array + offset, "x", color='red', markersize=5)

            if items_in_group:
                current_y_offset += (len(items_in_group) - 1) * intra_group_offset

        if self._is_zoomed and current_xlim is not None:
            ax.set_xlim(current_xlim)
            ax.set_ylim(current_ylim)
        else:
            if self.plot_settings["x_lim"]: ax.set_xlim(self.plot_settings["x_lim"])
            if self.plot_settings["y_lim"]: ax.set_ylim(self.plot_settings["y_lim"])

        # Configurar locator do eixo X com proteção contra muitos ticks
        if self.plot_settings["x_ticks"] and self.plot_settings["x_ticks"] >= 0.1:
            ax.xaxis.set_major_locator(MultipleLocator(self.plot_settings["x_ticks"]))
        else:
            # Usar MaxNLocator automático para eixo X
            ax.xaxis.set_major_locator(MaxNLocator(nbins=10, prune='both'))

        # Configurar locator do eixo Y com proteção contra muitos ticks
        if self.plot_settings["y_ticks"] and self.plot_settings["y_ticks"] >= 0.1:
            ax.yaxis.set_major_locator(MultipleLocator(self.plot_settings["y_ticks"]))
        else:
            # Usar MaxNLocator com limite de ticks para prevenir travamento
            # Isso previne o erro "Locator attempting to generate 53581 ticks"
            ax.yaxis.set_major_locator(MaxNLocator(nbins=10, prune='both'))

        ax.xaxis.set_visible(self.plot_settings["x_visible"])
        ax.yaxis.set_visible(self.plot_settings["y_visible"])

        if self.plot_settings.get("grid_visible", True):
            ax.grid(True, linestyle='--', alpha=0.6)
        else:
            ax.grid(False)

        ax.set_xlabel(self.plot_settings["x_label"], fontsize=self.plot_settings["x_label_fontsize"],
                      fontweight='bold' if self.plot_settings["x_label_bold"] else 'normal',
                      fontstyle='italic' if self.plot_settings["x_label_italic"] else 'normal')
        ax.set_ylabel(self.plot_settings["y_label"], fontsize=self.plot_settings["y_label_fontsize"],
                      fontweight='bold' if self.plot_settings["y_label_bold"] else 'normal',
                      fontstyle='italic' if self.plot_settings["y_label_italic"] else 'normal')

        line_width = self.plot_settings.get('axes_linewidth', 1.0)
        for spine in ax.spines.values():
            spine.set_linewidth(line_width)

        ax.tick_params(axis='x',
                       direction=self.plot_settings.get('xtick_direction', 'in'),
                       labelsize=self.plot_settings.get('xtick_labelsize', 10),
                       width=self.plot_settings.get('xtick_width', 0.8),
                       bottom=self.plot_settings.get('xtick_visible', True),
                       labelbottom=self.plot_settings.get('xtick_label_visible', True))

        ax.tick_params(axis='y',
                       direction=self.plot_settings.get('ytick_direction', 'in'),
                       labelsize=self.plot_settings.get('ytick_labelsize', 10),
                       width=self.plot_settings.get('ytick_width', 0.8),
                       left=self.plot_settings.get('ytick_visible', True),
                       labelleft=self.plot_settings.get('ytick_label_visible', True))

        # Configurar e exibir legenda com configurações personalizadas
        if any(line.get_label() for line in ax.get_lines()):
            self._update_legend()

        # IMPORTANTE: NÃO usar tight_layout quando há legenda com posição customizada
        # pois isso causa o redimensionamento indesejado do gráfico.
        # Se a legenda tiver bbox_to_anchor (posição customizada), não aplicar tight_layout
        has_custom_legend_position = (
            self._legend_ref and
            self.legend_settings.get("bbox_to_anchor") is not None
        )

        if not has_custom_legend_position:
            # Só aplicar tight_layout se não houver legenda com posição customizada
            try:
                self.plot_canvas.figure.tight_layout(rect=[0, 0, 1, 1])
            except Exception as e:
                logging.warning(f"Erro ao aplicar tight_layout: {e}")

        self.plot_canvas.draw()

        if not self._is_zoomed:
            self._original_x_lim = ax.get_xlim()
            self._original_y_lim = ax.get_ylim()

    def _update_legend(self):
        """Atualiza a legenda com as configurações atuais."""
        ax = self.plot_canvas.axes
        settings = self.legend_settings

        # Remover legenda antiga se existir
        if self._legend_ref:
            try:
                self._legend_ref.remove()
            except:
                pass

        # Obter handles e labels
        handles, labels = ax.get_legend_handles_labels()

        if not handles:
            return

        # Inverter ordem se necessário
        if settings.get("reverse", False):
            handles = handles[::-1]
            labels = labels[::-1]

        # Verificar se há posição bbox salva (coordenadas customizadas)
        bbox_to_anchor = settings.get("bbox_to_anchor", None)
        loc_setting = settings.get("loc", "best")

        # Criar nova legenda com todas as configurações
        if bbox_to_anchor:
            # Se há bbox customizado, usar ele
            # Usar bbox_transform=ax.transAxes para que a legenda seja posicionada
            # em relação ao axes, não à figura, evitando que o tight_layout empurre o gráfico
            # clip_on=False permite que a legenda fique fora dos limites do axes
            self._legend_ref = ax.legend(
                handles, labels,
                bbox_to_anchor=bbox_to_anchor,
                bbox_transform=ax.transAxes,  # IMPORTANTE: usar transformação do axes
                loc='upper left',  # Usar upper left como referência para bbox
                fontsize=settings.get("fontsize", 10),
                frameon=settings.get("frameon", True),
                fancybox=settings.get("fancybox", True),
                shadow=settings.get("shadow", False),
                framealpha=settings.get("framealpha", 0.9),
                ncol=settings.get("ncol", 1),
            )
            # Permitir que a legenda seja desenhada fora dos limites do axes
            self._legend_ref.set_clip_on(False)
        else:
            # Usar localização padrão
            # Se loc='best', o matplotlib pode colocar a legenda fora do gráfico
            # Para evitar isso, usamos bbox_to_anchor=(1, 1) com loc padrão
            # quando a localização é 'best' e não há bbox customizado
            if loc_setting == "best":
                # Para 'best', usar localização automática dentro do gráfico
                self._legend_ref = ax.legend(
                    handles, labels,
                    loc='best',
                    fontsize=settings.get("fontsize", 10),
                    frameon=settings.get("frameon", True),
                    fancybox=settings.get("fancybox", True),
                    shadow=settings.get("shadow", False),
                    framealpha=settings.get("framealpha", 0.9),
                    ncol=settings.get("ncol", 1),
                )
            else:
                # Para outras localizações, usar normalmente
                self._legend_ref = ax.legend(
                    handles, labels,
                    loc=loc_setting,
                    fontsize=settings.get("fontsize", 10),
                    frameon=settings.get("frameon", True),
                    fancybox=settings.get("fancybox", True),
                    shadow=settings.get("shadow", False),
                    framealpha=settings.get("framealpha", 0.9),
                    ncol=settings.get("ncol", 1),
                )

        # Aplicar estilo de fonte (negrito/itálico)
        for text in self._legend_ref.get_texts():
            text.set_fontweight(settings.get("fontweight", "normal"))
            text.set_fontstyle(settings.get("fontstyle", "normal"))

        # Configurar arrastar
        if settings.get("draggable", True):
            self._legend_ref.set_draggable(True)
            # Conectar evento de liberação do mouse para salvar posição
            self.plot_canvas.mpl_connect('button_release_event', self._on_legend_release)
        else:
            self._legend_ref.set_draggable(False)

        logging.debug(f"Legenda atualizada com configurações: {settings}")

    def _on_legend_release(self, event):
        """Chamado quando o botão do mouse é liberado. Salva a posição da legenda se foi arrastada."""
        if not self._legend_ref:
            return

        try:
            # Obter a posição atual da legenda em coordenadas relativas (0-1)
            # O matplotlib já ajusta isso automaticamente quando a legenda é arrastada
            bbox = self._legend_ref.get_window_extent()
            # Converter para coordenadas do axes (0-1)
            bbox_axes = bbox.transformed(self.plot_canvas.axes.transAxes.inverted())

            # Salvar como bbox_to_anchor (canto superior esquerdo da legenda)
            new_bbox = (bbox_axes.x0, bbox_axes.y1)

            # Só salva se a posição mudou significativamente
            current_bbox = self.legend_settings.get("bbox_to_anchor", None)
            if current_bbox is None or abs(new_bbox[0] - current_bbox[0]) > 0.01 or abs(new_bbox[1] - current_bbox[1]) > 0.01:
                self.legend_settings["bbox_to_anchor"] = new_bbox
                self.legend_settings["loc"] = "upper left"  # Usar upper left como referência

                # Permitir que a legenda fique fora dos limites sem recorte
                self._legend_ref.set_clip_on(False)

                self.set_dirty()  # Marca como alterado
                logging.debug(f"Posição da legenda salva: {new_bbox}")
        except Exception as e:
            logging.debug(f"Erro ao salvar posição da legenda: {e}")

    def open_legend_dialog(self):
        """Abre o dialog de configuração da legenda."""
        dialog = LegendDialog(self, self.legend_settings.copy())

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.legend_settings = dialog.get_settings()
            self._update_legend()
            self.plot_canvas.draw()
            self.set_dirty()
            logging.info("Configurações da legenda aplicadas")

    def apply_legend_settings(self, settings):
        """Aplica configurações da legenda sem fechar o dialog (botão Aplicar)."""
        self.legend_settings = settings
        self._update_legend()
        self.plot_canvas.draw()
        self.set_dirty()
        logging.debug("Configurações da legenda aplicadas (Aplicar)")

    def show_exp_list_context_menu(self, position):
        if not self.exp_list_widget.selectedItems(): return
        menu = QMenu(self)

        # Opção de normalização por pico específico (apenas se 1 item selecionado)
        selected_items = self.exp_list_widget.selectedItems()
        if len(selected_items) == 1:
            normalize_peak_action = menu.addAction("Normalizar por Pico Específico...")
            normalize_peak_action.triggered.connect(lambda: self.start_peak_selection_mode(
                selected_items[0].data(Qt.ItemDataRole.UserRole)
            ))
            menu.addSeparator()

        remove_action = menu.addAction("Remover da Lista")
        action = menu.exec(self.exp_list_widget.mapToGlobal(position))
        if action == remove_action:
            self._add_state_to_history()
            ids_to_remove = [it.data(Qt.ItemDataRole.UserRole) for it in selected_items]
            for item_id in ids_to_remove:
                self.remove_plot_item(item_id)

    def show_cif_list_context_menu(self, position):
        selected_list_items = self.cif_list_widget.selectedItems()
        if not selected_list_items: return

        menu = QMenu(self)

        if len(selected_list_items) == 1:
            item_id = selected_list_items[0].data(Qt.ItemDataRole.UserRole)
            plot_item = self._find_plot_item_by_id(item_id)

            # Adicionar opção "Reflexão" para itens CIF
            if plot_item and plot_item.get("type") == "cif":
                reflection_action = menu.addAction("Reflexão")
                reflection_action.triggered.connect(lambda: self.show_reflection_dialog(item_id))

                # Adicionar opção "Visualizar 3D" se o visualizador estiver disponível
                if hasattr(self, 'structure_viewer') and self.structure_viewer is not None:
                    view_3d_action = menu.addAction("Visualizar 3D")
                    view_3d_action.triggered.connect(lambda: self.view_structure_3d(item_id))

                menu.addSeparator()

            if plot_item and plot_item.get("type") == "cif" and plot_item.get("is_modified"):
                save_cif_action = menu.addAction("Salvar CIF")
                save_cif_action.triggered.connect(lambda: self.save_modified_cif(item_id))

                reset_action = menu.addAction("Resetar para Original")
                reset_action.triggered.connect(lambda: self.reset_cif_to_original(item_id))
                menu.addSeparator()

        remove_action = menu.addAction("Remover da Lista")
        remove_action.triggered.connect(self.remove_selected_cifs)

        menu.exec(self.cif_list_widget.mapToGlobal(position))

    def remove_selected_cifs(self):
        self._add_state_to_history()
        selected_items = self.cif_list_widget.selectedItems()
        ids_to_remove = [it.data(Qt.ItemDataRole.UserRole) for it in selected_items]
        for item_id in ids_to_remove:
            self.remove_plot_item(item_id)

    def show_reflection_dialog(self, item_id):
        """Abre o diálogo de reflexões para o item CIF selecionado."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("type") != "cif":
            QMessageBox.warning(self, ptr("Seleção Inválida"),
                              ptr("Por favor, selecione um item CIF válido."))
            return

        # Obter comprimento de onda
        try:
            wl_text = self.wavelength_combo.currentText()
            if wl_text == "Outro":
                wavelength = float(self.wavelength_custom_input.text())
            else:
                wavelength = float(wl_text.split('(')[1].split(' ')[0])
        except (ValueError, IndexError):
            QMessageBox.critical(self, ptr("Erro"),
                               ptr("Comprimento de onda inválido."))
            return

        # Obter 2θ máximo
        max_2theta = self.max_2theta_spin.value()

        # Abrir diálogo de reflexões
        dialog = ReflectionDialog(
            plot_item["cif_content"],
            wavelength,
            max_2theta,
            plot_item["label"],
            self
        )
        dialog.exec()

    def view_structure_3d(self, item_id):
        """Visualiza a estrutura cristalina 3D do CIF selecionado."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("type") != "cif":
            QMessageBox.warning(self, ptr("Seleção Inválida"),
                              ptr("Por favor, selecione um item CIF válido."))
            return

        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            QMessageBox.warning(self, ptr("Visualizador Indisponível"),
                              "O visualizador 3D não está disponível.\n"
                              "Certifique-se de que 'pyqtgraph' e 'PyOpenGL' estão instalados.")
            return

        # Salvar conteúdo CIF em arquivo temporário
        import tempfile
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.cif', delete=False, encoding='utf-8') as tmp_file:
                tmp_file.write(plot_item["cif_content"])
                tmp_cif_path = tmp_file.name

            # Carregar no visualizador 3D
            success = self.structure_viewer.load_cif(tmp_cif_path)

            # Remover arquivo temporário
            try:
                os.remove(tmp_cif_path)
            except:
                pass

            if success:
                # Mudar para a aba do visualizador 3D
                self.plot_tab_widget.setCurrentWidget(self.structure_viewer)
                logging.info(f"Estrutura 3D carregada para: {plot_item['label']}")

        except Exception as e:
            error_msg = f"Erro ao visualizar estrutura 3D: {str(e)}"
            logging.error(error_msg)
            QMessageBox.critical(self, ptr("Erro"), error_msg)

    def _on_cif_selection_changed(self):
        """Callback quando a seleção de CIF muda - marca para atualizar visualizador 3D."""
        # Verificar se o visualizador 3D está disponível
        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            return

        selected_items = self.cif_list_widget.selectedItems()

        # Se nenhum item selecionado, limpar visualizador
        if not selected_items:
            self._pending_3d_cif_id = None
            if hasattr(self, '_render_3d_button'):
                self._render_3d_button.setEnabled(False)
            return

        # Pegar apenas o primeiro item selecionado
        if len(selected_items) > 0:
            item_id = selected_items[0].data(Qt.ItemDataRole.UserRole)
            plot_item = self._find_plot_item_by_id(item_id)

            if not plot_item or plot_item.get("type") != "cif":
                self._pending_3d_cif_id = None
                if hasattr(self, '_render_3d_button'):
                    self._render_3d_button.setEnabled(False)
                return

            # Verificar se o CIF está calculado (tem dados)
            if plot_item.get("data") is None or len(plot_item.get("data", [])) == 0:
                # CIF não calculado
                self._pending_3d_cif_id = None

                # Limpar visualizador
                if hasattr(self.structure_viewer, '_clear_structure'):
                    self.structure_viewer._clear_structure()

                # Desativar botão renderizar
                if hasattr(self, '_render_3d_button'):
                    self._render_3d_button.setEnabled(False)

                # Mostrar aviso APENAS se estiver na aba do visualizador 3D
                if self.plot_tab_widget.currentWidget() == self.structure_viewer:
                    if item_id not in self._warned_uncalculated_cifs:
                        self._warned_uncalculated_cifs.add(item_id)
                        QMessageBox.information(
                            self,
                            ptr("CIF não calculado"),
                            f"O CIF '{plot_item['label']}' ainda não foi calculado.\n\n"
                            "Clique em 'Calcular Selecionado(s)' para gerar o difratograma e "
                            "habilitar a visualização 3D."
                        )
                return

            # CIF calculado - marcar como pendente e ativar botão
            self._pending_3d_cif_id = item_id
            if hasattr(self, '_render_3d_button'):
                self._render_3d_button.setEnabled(True)
                self._render_3d_button.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")

            # IMPORTANTE: NÃO renderizar automaticamente ao selecionar
            # O usuário deve clicar explicitamente em "Visualizar Estrutura"
            logging.debug(f"CIF selecionado: {plot_item['label']} - aguardando clique em 'Visualizar Estrutura'")

    def _load_cif_to_3d_viewer(self, item_id):
        """Carrega um CIF no visualizador 3D (método interno)."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("type") != "cif":
            return

        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            return

        # Salvar conteúdo CIF em arquivo temporário
        import tempfile
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.cif', delete=False, encoding='utf-8') as tmp_file:
                tmp_file.write(plot_item["cif_content"])
                tmp_cif_path = tmp_file.name

            # Carregar no visualizador 3D
            success = self.structure_viewer.load_cif(tmp_cif_path)

            # Remover arquivo temporário
            try:
                os.remove(tmp_cif_path)
            except:
                pass

            if success:
                logging.info(f"Estrutura 3D atualizada para: {plot_item['label']}")

                # Desmarcar mudanças pendentes
                self._has_pending_3d_changes = False

                # Desativar botão renderizar
                if hasattr(self, '_render_3d_button') and self._render_3d_button:
                    self._render_3d_button.setEnabled(False)
                    self._render_3d_button.setStyleSheet("")

        except Exception as e:
            logging.error(f"Erro ao atualizar visualizador 3D: {e}")

    def _on_plot_tab_changed(self, index):
        """
        Callback quando o usuário troca de aba (Difratograma <-> Visualizador 3D).
        OTIMIZAÇÃO CRÍTICA: Suspende completamente o widget inativo para evitar travamentos.
        """
        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            return

        # Detectar qual aba está ativa
        is_3d_tab_active = (index == 1)  # 0=Difratograma, 1=Estrutura 3D

        if is_3d_tab_active:
            # ========== ATIVAR VISUALIZADOR 3D ==========
            self.structure_viewer.setEnabled(True)
            if hasattr(self.structure_viewer, 'view_widget'):
                self.structure_viewer.view_widget.setEnabled(True)
                # Desbloquear eventos de mouse/teclado
                self.structure_viewer.view_widget.blockSignals(False)

            # Reativar timer do gizmo se estava ativo
            if hasattr(self.structure_viewer, '_gizmo_sync_timer'):
                if hasattr(self.structure_viewer, '_gizmo_timer_was_active') and self.structure_viewer._gizmo_timer_was_active:
                    if self.structure_viewer._gizmo_sync_timer and not self.structure_viewer._gizmo_sync_timer.isActive():
                        self.structure_viewer._gizmo_sync_timer.start(50)
                    self.structure_viewer._gizmo_timer_was_active = False

            # ========== SUSPENDER MATPLOTLIB (DIFRATOGRAMA) ==========
            if hasattr(self, 'plot_canvas'):
                # Bloquear completamente atualizações do matplotlib
                self.plot_canvas.setUpdatesEnabled(False)
                # Bloquear eventos de mouse/teclado no canvas
                self.plot_canvas.setEnabled(False)

            logging.debug("✅ [TAB CHANGE] Visualizador 3D ATIVO | Matplotlib SUSPENSO")

        else:
            # ========== ATIVAR MATPLOTLIB (DIFRATOGRAMA) ==========
            if hasattr(self, 'plot_canvas'):
                self.plot_canvas.setEnabled(True)
                self.plot_canvas.setUpdatesEnabled(True)
                # Forçar redesenho apenas se houver mudanças pendentes
                if hasattr(self, '_has_pending_plot_updates') and self._has_pending_plot_updates:
                    self.plot_canvas.draw_idle()
                    self._has_pending_plot_updates = False

            # ========== SUSPENDER VISUALIZADOR 3D ==========
            # Parar todas as animações em execução
            if hasattr(self.structure_viewer, 'stop_animation'):
                self.structure_viewer.stop_animation()

            # Pausar timer de sincronização do gizmo (economizar CPU)
            if hasattr(self.structure_viewer, '_gizmo_sync_timer'):
                if self.structure_viewer._gizmo_sync_timer and self.structure_viewer._gizmo_sync_timer.isActive():
                    self.structure_viewer._gizmo_sync_timer.stop()
                    self.structure_viewer._gizmo_timer_was_active = True

            # Desabilitar renderização OpenGL
            if hasattr(self.structure_viewer, 'view_widget'):
                self.structure_viewer.view_widget.setEnabled(False)
                # Bloquear eventos de mouse/teclado no visualizador
                self.structure_viewer.view_widget.blockSignals(True)

            logging.debug("✅ [TAB CHANGE] Matplotlib ATIVO | Visualizador 3D SUSPENSO")

    def _render_3d_structure(self):
        """Renderiza/atualiza a estrutura 3D do CIF selecionado."""
        if self._pending_3d_cif_id is None:
            return

        plot_item = self._find_plot_item_by_id(self._pending_3d_cif_id)
        if not plot_item:
            return

        # Verificar se está calculado
        if plot_item.get("data") is None or len(plot_item.get("data", [])) == 0:
            # Mostrar aviso apenas se estiver na aba do visualizador
            if self.plot_tab_widget.currentWidget() == self.structure_viewer or \
               (hasattr(self.plot_tab_widget.currentWidget(), 'children') and
                self.structure_viewer in self.plot_tab_widget.currentWidget().children()):
                if self._pending_3d_cif_id not in self._warned_uncalculated_cifs:
                    self._warned_uncalculated_cifs.add(self._pending_3d_cif_id)
                    QMessageBox.information(
                        self,
                        ptr("CIF não calculado"),
                        f"O CIF '{plot_item['label']}' ainda não foi calculado.\n\n"
                        "Clique em 'Calcular Selecionado(s)' para gerar o difratograma e "
                        "habilitar a visualização 3D."
                    )
            return

        # Carregar no visualizador
        self._load_cif_to_3d_viewer(self._pending_3d_cif_id)

        logging.info(f"Estrutura 3D renderizada para: {plot_item['label']}")

    def _open_3d_tools_dialog(self):
        """Abre diálogo de configurações do visualizador 3D."""
        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            return

        dialog = Viewer3DToolsDialog(self.structure_viewer, self)
        dialog.exec()

    def _open_animation_dialog(self):
        """Abre diálogo de animações 3D."""
        if not hasattr(self, 'structure_viewer') or self.structure_viewer is None:
            return

        if not self.structure_viewer.structure:
            QMessageBox.information(
                self,
                ptr("Sem Estrutura"),
                ptr("Carregue uma estrutura CIF primeiro para usar as animações.")
            )
            return

        dialog = Viewer3DAnimationDialog(self.structure_viewer, self)
        dialog.exec()

    def save_modified_cif(self, item_id):
        """Salva o CIF modificado pelo editor em um arquivo .cif escolhido pelo usuário."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or not plot_item.get("is_modified"):
            return

        cif_content = plot_item.get("cif_content", "")
        if not cif_content:
            QMessageBox.warning(self, ptr("Erro"), ptr("Não há conteúdo CIF para salvar."))
            return

        # Nome padrão: nome original + _PDRX-m
        original_label = plot_item.get("label", "CIF")
        default_name = f"{original_label}_PDRX-m.cif"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar CIF Modificado",
            default_name,
            "Arquivo CIF (*.cif);;Todos os arquivos (*.*)"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(cif_content)
            logging.info(f"CIF modificado salvo em: {file_path}")
            QMessageBox.information(self, ptr("Salvo"), f"CIF modificado salvo com sucesso em:\n{file_path}")
        except Exception as e:
            logging.error(f"Erro ao salvar CIF modificado: {e}")
            QMessageBox.critical(self, ptr("Erro"), f"Não foi possível salvar o arquivo:\n{e}")

    def reset_cif_to_original(self, item_id):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or not plot_item.get("is_modified"):
            return

        reply = QMessageBox.question(self, ptr("Resetar CIF"),
                                     f"Tem certeza que deseja resetar '{plot_item['label']}' para o seu estado original?\n"
                                     "Todas as modificações de rede e parâmetros de simulação serão perdidos.",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No:
            return

        self._add_state_to_history()

        if self.project_directory and plot_item.get('modified_cif_path'):
            path_to_delete = os.path.join(self.project_directory, plot_item['modified_cif_path'])
            if os.path.exists(path_to_delete):
                try:
                    os.remove(path_to_delete)
                    logging.info(f"Arquivo CIF modificado '{path_to_delete}' removido.")
                except OSError as e:
                    QMessageBox.warning(self, ptr("Erro ao Apagar"), f"Não foi possível apagar o arquivo modificado:\n{e}")
                    logging.error(f"Erro ao apagar arquivo modificado '{path_to_delete}': {e}")

        plot_item['cif_content'] = plot_item.get('original_cif_content', '')
        plot_item['is_modified'] = False
        plot_item['simulation_params'] = {}
        plot_item.pop('modified_cif_path', None)

        self.run_simulation_for_item(plot_item)
        self._repopulate_all_lists()
        self.set_dirty()
        QMessageBox.information(self, ptr("Resetado"), f"'{plot_item['label']}' foi resetado para o original.")

    def remove_plot_item(self, item_id_to_remove):
        logging.info(f"🗑 Removendo item: {item_id_to_remove}")

        # Verificar se o item sendo removido é um CIF
        plot_item = self._find_plot_item_by_id(item_id_to_remove)
        is_cif_being_removed = plot_item and plot_item.get("type") in ["cif", "multiphase"]

        # Se for CIF E houver visualizador 3D, limpar estrutura
        if is_cif_being_removed:
            if hasattr(self, 'structure_viewer') and self.structure_viewer is not None:
                try:
                    logging.info("✓ CIF removido - limpando estrutura 3D existente")

                    # Limpar estrutura 3D
                    if hasattr(self.structure_viewer, '_clear_structure'):
                        self.structure_viewer._clear_structure()

                    # Resetar estado da estrutura
                    self.structure_viewer.structure = None

                    # Atualizar label de informações
                    if hasattr(self.structure_viewer, 'info_label'):
                        self.structure_viewer.info_label.setText(ptr("Nenhuma estrutura carregada"))
                        self.structure_viewer.info_label.setStyleSheet("color: gray; font-style: italic;")

                    # Limpar legenda de elementos
                    if hasattr(self.structure_viewer, '_element_legend_items'):
                        for label_widget in self.structure_viewer._element_legend_items:
                            label_widget.setVisible(False)
                        self.structure_viewer._element_legend_items.clear()

                    logging.debug("✓ Estrutura 3D completamente limpa")

                except Exception as e:
                    logging.error(f"Erro ao limpar estrutura 3D: {e}")

            # Resetar estado de pendências
            if hasattr(self, '_pending_3d_cif_id') and self._pending_3d_cif_id == item_id_to_remove:
                self._pending_3d_cif_id = None
                self._has_pending_3d_changes = False

            # Desativar botão de renderização
            if hasattr(self, '_render_3d_button'):
                self._render_3d_button.setEnabled(False)
                self._render_3d_button.setStyleSheet("")

        # Remover o item da lista
        self.plot_items = [p for p in self.plot_items if p["id"] != item_id_to_remove]
        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def open_plot_customization_dialog(self):
        dialog = PlotCustomizationDialog(self.plot_canvas.axes, self.plot_settings, self.plot_items,
                                         self.group_settings, self)
        dialog.settings_applied.connect(self.apply_plot_customizations)
        dialog.exec()

    def apply_plot_customizations(self, axis_settings, line_settings, group_spacing):
        self._add_state_to_history()
        self.plot_settings.update(axis_settings)
        self.group_settings["spacing"] = group_spacing
        for settings in line_settings:
            item = self._find_plot_item_by_id(settings["id"])
            if item:
                item.update(settings)
        self._is_zoomed = False
        self._repopulate_all_lists()
        self._redraw_all_plots()
        self.set_dirty()

    def open_smooth_dialog(self):
        selected_items = self.exp_list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, ptr("Sem Dados"),
                                ptr("Selecione um ou mais arquivos experimentais da lista para suavizar."))
            return
        if not SCIPY_AVAILABLE:
            QMessageBox.critical(self, ptr("Dependência Faltando"), ptr("A biblioteca 'scipy' é necessária para a suavização."))
            return

        dialog = SmoothDialog(self)
        if dialog.exec():
            self._add_state_to_history()
            settings = dialog.get_settings()
            for list_item in selected_items:
                item_id = list_item.data(Qt.ItemDataRole.UserRole)
                self.apply_smoothing(item_id, settings["window"], settings["order"])

    def apply_smoothing(self, item_id, window_length, polyorder):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("data") is None: return

        try:
            _, y_data = plot_item["data"]
            smoothed_y = smooth_data(y_data, window_length, polyorder)
            plot_item["data"][1] = smoothed_y
            plot_item["is_smoothed"] = True
            if "analysis" in plot_item: plot_item["analysis"]["peaks"] = None
            self._redraw_all_plots()
            self.set_dirty()
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro na Suavização"), f"Ocorreu um erro: {e}")

    def open_peak_detect_dialog(self):
        selected_items = self.exp_list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, ptr("Sem Dados"),
                                ptr("Selecione um ou mais arquivos experimentais da lista para detectar picos."))
            return
        if not SCIPY_AVAILABLE:
            QMessageBox.critical(self, ptr("Dependência Faltando"),
                                 ptr("A biblioteca 'scipy' é necessária para a detecção de picos."))
            return

        dialog = PeakDetectDialog(self)
        if dialog.exec():
            self._add_state_to_history()
            settings = dialog.get_settings()
            for list_item in selected_items:
                item_id = list_item.data(Qt.ItemDataRole.UserRole)
                self.apply_peak_detection(item_id, settings["prominence"], settings["width"])

    def apply_peak_detection(self, item_id, prominence, width):
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("data") is None: return

        try:
            x_data, y_data = plot_item["data"]
            peak_positions, peak_intensities = detect_peaks(y_data, x_data, prominence, width)
            if "analysis" not in plot_item: plot_item["analysis"] = {}
            plot_item["analysis"]["peaks"] = (peak_positions.tolist(), peak_intensities.tolist())
            self._redraw_all_plots()
            self.set_dirty()
        except Exception as e:
            QMessageBox.critical(self, ptr("Erro na Detecção de Picos"), f"Ocorreu um erro: {e}")

    def open_save_plot_dialog(self):
        default_path = self.project_directory if self.project_directory else ""
        dialog = SavePlotDialog(default_path=default_path, parent=self)
        if dialog.exec():
            settings = dialog.get_settings()
            try:
                # NÃO usar bbox_inches='tight' quando há legenda customizada
                # pois isso causaria redimensionamento indesejado
                has_custom_legend = self.legend_settings.get("bbox_to_anchor") is not None

                if has_custom_legend:
                    # Salvar sem bbox_inches para manter layout exato
                    self.plot_canvas.figure.savefig(
                        settings["path"],
                        dpi=settings["dpi"],
                        format=settings["format"]
                    )
                else:
                    # Usar bbox_inches='tight' apenas se não houver legenda customizada
                    self.plot_canvas.figure.savefig(
                        settings["path"],
                        dpi=settings["dpi"],
                        format=settings["format"],
                        bbox_inches='tight'
                    )

                QMessageBox.information(self, ptr("Sucesso"), f"Gráfico salvo em:\n{settings['path']}")
            except Exception as e:
                QMessageBox.critical(self, ptr("Erro ao Salvar"), f"Não foi possível salvar o gráfico:\n{e}")

    # --- ALTERAÇÃO DE FUNCIONALIDADE: Adicionar Função Exportar para Origin ---
    def export_to_xlsx_for_origin(self):
        """
        Exporta todos os dados atualmente carregados para um
        único ficheiro .xlsx, onde cada gráfico é uma aba.
        """
        
        # 1. Verifica se a biblioteca está disponível
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, ptr("Dependência Faltando"),
                                 "A biblioteca 'openpyxl' é necessária para exportar para Excel.\n"
                                 "Por favor, instale-a com: pip install openpyxl")
            return
            
        # 2. Verifica se há dados
        if not self.plot_items:
            QMessageBox.warning(self,ptr("Exportar para Excel"),
                                ptr("Não há dados carregados para exportar."))
            return
            
        # 3. Pergunta ao utilizador onde salvar
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar para Origin (.xlsx)",
            self.project_directory if self.project_directory else "",
            "Pasta de Trabalho do Excel (*.xlsx)"
        )

        if not filepath:
            return # Utilizador cancelou

        # 4. Cria o ficheiro Excel
        try:
            wb = openpyxl.Workbook()
            
            # Remove a aba padrão "Sheet"
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # 5. Itera sobre os items visíveis (experimentais e teóricos)
            items_to_export = [item for item in self.plot_items if item.get("data") is not None]

            if not items_to_export:
                QMessageBox.warning(self, ptr("Sem Dados"), ptr("Nenhum dos itens carregados possui dados simulados ou lidos."))
                return

            for item in items_to_export:
                # O nome da aba tem restrições de caracteres e tamanho
                sheet_name = item["label"].replace('[','').replace(']','').replace('*','').replace('?','').replace(':','').replace('/','\\')
                sheet_name = sheet_name[:31] # Limite de 31 caracteres do Excel
                
                # Evita nomes duplicados
                if sheet_name in wb.sheetnames:
                    sheet_name = sheet_name[:28] + "_1"
                
                ws = wb.create_sheet(title=sheet_name)
                
                # 6. Escreve Cabeçalhos
                ws['A1'] = "2-Theta"
                ws['B1'] = "Intensidade"
                
                # 7. Escreve os Dados
                x_data, y_data = item["data"]
                for i, (x, y) in enumerate(zip(x_data, y_data), start=2):
                    ws.cell(row=i, column=1, value=x)
                    ws.cell(row=i, column=2, value=y)
            
            # 8. Salva o ficheiro
            wb.save(filepath)
            QMessageBox.information(self, ptr("Exportação Concluída"),
                                    f"Dados exportados com sucesso para:\n{filepath}\n\n"
                                    "Arraste este ficheiro para o Origin para importar.")

        except Exception as e:
            QMessageBox.critical(self, ptr("Erro na Exportação"),
                                 f"Ocorreu um erro ao exportar para .xlsx:\n{e}")
            logging.exception("Erro ao exportar para XLSX:")
    # --- FIM DA ALTERAÇÃO ---

    def _find_plot_item_by_id(self, item_id):
        return next((p for p in self.plot_items if p["id"] == item_id), None)

    def _get_next_color(self):
        used_colors = {item.get("color") for item in self.plot_items if item.get("color")}

        for i in range(len(DEFAULT_PLOT_COLORS)):
            next_color_index = (self.color_cycler_index + i) % len(DEFAULT_PLOT_COLORS)
            next_color = DEFAULT_PLOT_COLORS[next_color_index]
            if next_color not in used_colors:
                self.color_cycler_index = next_color_index + 1
                return next_color

        color = DEFAULT_PLOT_COLORS[self.color_cycler_index % len(DEFAULT_PLOT_COLORS)]
        self.color_cycler_index += 1
        return color

    def on_wavelength_change(self, index):
        self.wavelength_custom_input.setVisible(self.wavelength_combo.currentText() == "Outro")

    def _add_state_to_history(self):
        state = {
            "plot_items": copy.deepcopy(self.plot_items),
            "plot_settings": copy.deepcopy(self.plot_settings),
            "group_settings": copy.deepcopy(self.group_settings)
        }
        self.history_stack = self.history_stack[:self.history_index + 1]
        self.history_stack.append(state)
        self.history_index += 1
        self._update_undo_redo_buttons()

    def _restore_state_from_history(self):
        if not (0 <= self.history_index < len(self.history_stack)):
            return

        state = self.history_stack[self.history_index]
        self.plot_items = copy.deepcopy(state["plot_items"])
        self.plot_settings = copy.deepcopy(state["plot_settings"])
        self.group_settings = copy.deepcopy(state["group_settings"])

        self.num_groups_spin.setValue(self.group_settings.get("count", 1))
        self._is_zoomed = False

        self._repopulate_all_lists()
        self._redraw_all_plots()
        self._update_undo_redo_buttons()
        self.set_dirty()

    def start_peak_selection_mode(self, item_id):
        """Inicia o modo de seleção de pico para normalização."""
        plot_item = self._find_plot_item_by_id(item_id)
        if not plot_item or plot_item.get("data") is None:
            QMessageBox.warning(self, ptr("Erro"), ptr("Não foi possível encontrar os dados do arquivo selecionado."))
            return

        # Ativar modo de seleção
        self._peak_selection_mode = True
        self._peak_selection_item_id = item_id

        # Mudar cursor para crosshair
        from PySide6.QtGui import QCursor
        self.plot_canvas.setCursor(QCursor(Qt.CursorShape.CrossCursor))

        # Desabilitar zoom selector temporariamente
        if self.zoom_selector:
            self.zoom_selector.set_active(False)

        # Mostrar mensagem de instrução
        QMessageBox.information(
            self,
            ptr("Seleção de Pico"),
            f"Modo de seleção de pico ativado para: {plot_item['label']}\n\n"
            "Clique no gráfico na posição do pico de interesse.\n"
            "O programa irá buscar o máximo na região próxima ao clique.\n\n"
            "Pressione ESC para cancelar."
        )

    def _process_peak_selection(self, peak_x):
        """Processa a seleção de pico e realiza a normalização."""
        # Desativar modo de seleção
        self._peak_selection_mode = False
        from PySide6.QtGui import QCursor
        self.plot_canvas.setCursor(QCursor(Qt.CursorShape.ArrowCursor))

        # Reativar zoom selector
        if self.zoom_selector:
            self.zoom_selector.set_active(True)

        # Buscar o item selecionado
        plot_item = self._find_plot_item_by_id(self._peak_selection_item_id)
        if not plot_item or plot_item.get("data") is None:
            QMessageBox.critical(self, ptr("Erro"), ptr("Dados não encontrados."))
            return

        x_data, y_data = plot_item["data"]

        try:
            # Normalizar pelo pico
            y_normalized, peak_intensity = normalize_by_peak(x_data, y_data, peak_x, window_size=0.5)

            # Contar outros experimentais
            other_experimentals = [
                item for item in self.plot_items
                if item.get("type") == "experimental"
                and item["id"] != self._peak_selection_item_id
                and item.get("data") is not None
            ]

            # Abrir diálogo de confirmação
            dialog = NormalizeByPeakConfirmDialog(
                plot_item["label"],
                peak_x,
                len(other_experimentals),
                self
            )

            if dialog.exec() != QDialog.DialogCode.Accepted:
                return  # Usuário cancelou

            # Adicionar ao histórico
            self._add_state_to_history()

            # Aplicar normalização ao item selecionado
            plot_item["data"][1] = y_normalized
            success_count = 1

            # Se usuário escolheu normalizar todos
            if dialog.should_normalize_others():
                for other_item in other_experimentals:
                    try:
                        other_x, other_y = other_item["data"]
                        other_y_norm, _ = normalize_by_peak(other_x, other_y, peak_x, window_size=0.5)
                        other_item["data"][1] = other_y_norm
                        success_count += 1
                    except Exception as e:
                        QMessageBox.warning(
                            self,
                            ptr("Aviso"),
                            f"Não foi possível normalizar '{other_item['label']}':\n{e}"
                        )

            # Redesenhar e atualizar
            self._redraw_all_plots()
            self.set_dirty()

            # Mensagem de sucesso
            QMessageBox.information(
                self,
                ptr("Normalização Concluída"),
                f"{success_count} arquivo(s) normalizado(s) com sucesso!\n\n"
                f"Pico de referência: 2θ = {peak_x:.3f}°\n"
                f"Intensidade do pico: {peak_intensity:.2f}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                ptr("Erro na Normalização"),
                f"Não foi possível normalizar pelo pico:\n{e}"
            )

    def undo_action(self):
        if self.history_index > 0:
            self.history_index -= 1
            self._restore_state_from_history()

    def redo_action(self):
        if self.history_index < len(self.history_stack) - 1:
            self.history_index += 1
            self._restore_state_from_history()

    def _update_undo_redo_buttons(self):
        self.undo_button.setEnabled(self.history_index > 0)
        self.redo_button.setEnabled(self.history_index < len(self.history_stack) - 1)


class PlotCustomizationDialog(QDialog):
    settings_applied = Signal(dict, list, dict)

    def __init__(self, axes, current_settings, plot_data, group_settings, parent=None):
        super().__init__(parent)
        self.setWindowTitle(ptr("Personalizar Gráfico"))
        self.setFixedWidth(520)
        self.resize(520, 700)
        set_polvo_icon(self)

        self.axes = axes
        self.current_settings = current_settings
        self.plot_data = plot_data
        self.group_settings = group_settings
        self.intra_group_spinboxes = {}

        self._init_ui()
        self._populate_initial_values()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        tabs = QTabWidget()
        layout.addWidget(tabs)

        scale_tab = QWidget()
        tabs.addTab(scale_tab, ptr("Escala e Rótulos"))
        main_scale_layout = QVBoxLayout(scale_tab)

        scroll_area_scale = QScrollArea()
        scroll_area_scale.setWidgetResizable(True)
        scroll_content_scale = QWidget()
        scale_layout = QVBoxLayout(scroll_content_scale)

        labels_group = QGroupBox(ptr("Rótulos dos Eixos"));
        labels_form = QFormLayout(labels_group)
        self.x_label_edit = QLineEdit();
        self.y_label_edit = QLineEdit()
        labels_form.addRow("Eixo X:", self.x_label_edit);
        labels_form.addRow("Eixo Y:", self.y_label_edit)
        scale_layout.addWidget(labels_group)
        x_group = QGroupBox(ptr("Eixo X"));
        x_form = QFormLayout(x_group)
        self.x_from = CustomDoubleSpinBox(wheel_step=0.1);
        self.x_from.setDecimals(4)
        self.x_from.setSingleStep(0.1)
        self.x_to = CustomDoubleSpinBox(wheel_step=0.1);
        self.x_to.setDecimals(4)
        self.x_to.setSingleStep(0.1)
        self.x_ticks = CustomDoubleSpinBox(wheel_step=0.1)
        self.x_ticks.setDecimals(4)
        self.x_ticks.setSingleStep(0.1)
        self.x_ticks.setMinimum(0.0)  # 0 = automático
        self.x_ticks.setMaximum(2000)
        self.x_ticks.setToolTip(
            "Intervalo entre marcações principais no eixo X.\n"
            "0 = Automático (recomendado para 2θ)\n"
            "Valores muito pequenos (< 0.1) podem causar problemas."
        )
        self.x_label_size = QSpinBox();
        self.x_label_bold = QCheckBox(ptr("Negrito"));
        self.x_label_italic = QCheckBox(ptr("Itálico"))
        self.x_visible = QCheckBox(ptr("Mostrar Eixo X"))
        x_form.addRow("De:", self.x_from);
        x_form.addRow("Para:", self.x_to);
        x_form.addRow("Major Ticks:", self.x_ticks)
        x_font_layout = QHBoxLayout();
        x_font_layout.addWidget(self.x_label_size);
        x_font_layout.addWidget(self.x_label_bold);
        x_font_layout.addWidget(self.x_label_italic)
        x_form.addRow("Fonte do Rótulo:", x_font_layout);
        x_form.addRow(self.x_visible)
        scale_layout.addWidget(x_group)
        y_group = QGroupBox(ptr("Eixo Y"));
        y_form = QFormLayout(y_group)
        self.y_from = CustomDoubleSpinBox(wheel_step=0.1);
        self.y_from.setDecimals(4)
        self.y_from.setSingleStep(0.1)
        self.y_to = CustomDoubleSpinBox(wheel_step=0.1);
        self.y_to.setDecimals(4)
        self.y_to.setSingleStep(0.1)
        self.y_ticks = CustomDoubleSpinBox(wheel_step=0.1)
        self.y_ticks.setDecimals(4)
        self.y_ticks.setSingleStep(0.1)
        self.y_ticks.setMinimum(0.0)  # 0 = automático
        self.y_ticks.setMaximum(2000)
        self.y_ticks.setToolTip(
            "Intervalo entre marcações principais no eixo Y.\n"
            "0 = Automático (recomendado)\n"
            "Valores muito pequenos (< 0.1) podem causar problemas."
        )
        self.y_label_size = QSpinBox();
        self.y_label_bold = QCheckBox(ptr("Negrito"));
        self.y_label_italic = QCheckBox(ptr("Itálico"))
        self.y_visible = QCheckBox(ptr("Mostrar Eixo Y"))
        y_form.addRow("De:", self.y_from);
        y_form.addRow("Para:", self.y_to);
        y_form.addRow("Major Ticks:", self.y_ticks)
        y_font_layout = QHBoxLayout();
        y_font_layout.addWidget(self.y_label_size);
        y_font_layout.addWidget(self.y_label_bold);
        y_font_layout.addWidget(self.y_label_italic)
        y_form.addRow("Fonte do Rótulo:", y_font_layout);
        y_form.addRow(self.y_visible)
        scale_layout.addWidget(y_group)
        self.grid_visible = QCheckBox(ptr("Mostrar Grade"));
        scale_layout.addWidget(self.grid_visible)

        stack_group = QGroupBox(ptr("Empilhamento de Gráficos"))
        stack_form = QFormLayout(stack_group)
        self.global_offset_spin = CustomDoubleSpinBox(wheel_step=0.05)
        self.global_offset_spin.setDecimals(3)
        self.global_offset_spin.setSingleStep(0.05)  # <-- CORREÇÃO: Adicionado passo fino
        self.global_offset_spin.setRange(-100, 100)  # <-- BÔNUS: Adicionado um range razoável

        self.inter_group_offset_spin = QDoubleSpinBox()
        self.inter_group_offset_spin.setDecimals(3)
        self.inter_group_offset_spin.setSingleStep(0.05)  # <-- CORREÇÃO: Adicionado passo fino
        self.inter_group_offset_spin.setRange(-100, 100)  # <-- BÔNUS: Adicionado um range razoável

        stack_form.addRow("Deslocamento Global:", self.global_offset_spin)
        stack_form.addRow("Espaçamento entre Grupos:", self.inter_group_offset_spin)

        num_groups = self.group_settings.get("count", 1)
        for i in range(num_groups):
            spinbox = QDoubleSpinBox()
            spinbox.setDecimals(3)
            spinbox.setSingleStep(0.05)  # <-- CORREÇÃO: Adicionado passo fino
            spinbox.setRange(-100, 100)  # <-- BÔNUS: Adicionado um range razoável
            spinbox.setValue(self.group_settings.get("spacing", {}).get(str(i), 0.0))
            stack_form.addRow(f"Espaçamento no Grupo {i + 1}:", spinbox)
            self.intra_group_spinboxes[i] = spinbox
        scale_layout.addWidget(stack_group)
        scale_layout.addStretch()

        scroll_area_scale.setWidget(scroll_content_scale)
        main_scale_layout.addWidget(scroll_area_scale)

        lines_tab = QWidget()
        tabs.addTab(lines_tab, ptr("Linhas e Legendas"))
        lines_layout = QVBoxLayout(lines_tab)

        palette_group = QGroupBox(ptr("Esquemas de Cores (Paletas)"))
        palette_layout = QHBoxLayout(palette_group)
        palette_layout.addWidget(QLabel(ptr("Aplicar paleta:")))
        self.palette_combo = QComboBox()
        self.palette_combo.addItems([
            "Padrão MatFinder", "Viridis (Perceptual)", "Plasma (Perceptual)",
            "Tab10 (Qualitativa)", "Set1 (Qualitativa)", "Paired (Qualitativa)"
        ])
        self.apply_palette_button = QPushButton(ptr("Aplicar Paleta"))
        palette_layout.addWidget(self.palette_combo)
        palette_layout.addWidget(self.apply_palette_button)
        lines_layout.addWidget(palette_group)

        self.lines_table = QTableWidget()
        self.lines_table.setColumnCount(5)
        self.lines_table.setHorizontalHeaderLabels(["Legenda", "Cor", "Estilo", "Espessura", "Visível"])
        lines_layout.addWidget(self.lines_table)

        axes_tab = QWidget()
        tabs.addTab(axes_tab, ptr("Aparência dos Eixos"))
        main_axes_layout = QVBoxLayout(axes_tab)

        scroll_area_axes = QScrollArea()
        scroll_area_axes.setWidgetResizable(True)
        scroll_content_axes = QWidget()
        axes_layout = QVBoxLayout(scroll_content_axes)

        axes_appearance_group = QGroupBox(ptr("Aparência Geral dos Eixos"))
        axes_appearance_form = QFormLayout(axes_appearance_group)
        self.axes_linewidth_spin = QDoubleSpinBox()
        self.axes_linewidth_spin.setRange(0.5, 5.0)
        self.axes_linewidth_spin.setSingleStep(0.1)
        self.axes_linewidth_spin.setValue(1.0)
        axes_appearance_form.addRow("Espessura das Linhas dos Eixos:", self.axes_linewidth_spin)
        axes_layout.addWidget(axes_appearance_group)

        xticks_group = QGroupBox(ptr("Marcadores do Eixo X (Ticks)"))
        xticks_form = QFormLayout(xticks_group)
        self.xtick_direction_combo = QComboBox()
        self.xtick_direction_combo.addItems(["Para Dentro", "Para Fora", "Sem Ticks"])
        self.xtick_labelsize_spin = QSpinBox()
        self.xtick_labelsize_spin.setRange(6, 20)
        self.xtick_width_spin = QDoubleSpinBox()
        self.xtick_width_spin.setRange(0.5, 5.0)
        self.xtick_width_spin.setSingleStep(0.1)
        self.xtick_visible_check = QCheckBox(ptr("Mostrar Marcadores (ticks)"))
        self.xtick_label_visible_check = QCheckBox(ptr("Mostrar Números"))
        xticks_form.addRow("Direção:", self.xtick_direction_combo)
        xticks_form.addRow("Espessura:", self.xtick_width_spin)
        xticks_form.addRow("Tamanho da Fonte:", self.xtick_labelsize_spin)
        xticks_form.addRow(self.xtick_visible_check)
        xticks_form.addRow(self.xtick_label_visible_check)
        axes_layout.addWidget(xticks_group)

        yticks_group = QGroupBox(ptr("Marcadores do Eixo Y (Ticks)"))
        yticks_form = QFormLayout(yticks_group)
        self.ytick_direction_combo = QComboBox()
        self.ytick_direction_combo.addItems(["Para Dentro", "Para Fora", "Sem Ticks"])
        self.ytick_labelsize_spin = QSpinBox()
        self.ytick_labelsize_spin.setRange(6, 20)
        self.ytick_width_spin = QDoubleSpinBox()
        self.ytick_width_spin.setRange(0.5, 5.0)
        self.ytick_width_spin.setSingleStep(0.1)
        self.ytick_visible_check = QCheckBox(ptr("Mostrar Marcadores (ticks)"))
        self.ytick_label_visible_check = QCheckBox(ptr("Mostrar Números"))
        yticks_form.addRow("Direção:", self.ytick_direction_combo)
        yticks_form.addRow("Espessura:", self.ytick_width_spin)
        yticks_form.addRow("Tamanho da Fonte:", self.ytick_labelsize_spin)
        yticks_form.addRow(self.ytick_visible_check)
        yticks_form.addRow(self.ytick_label_visible_check)
        axes_layout.addWidget(yticks_group)
        axes_layout.addStretch()

        scroll_area_axes.setWidget(scroll_content_axes)
        main_axes_layout.addWidget(scroll_area_axes)

        self.button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.clicked.connect(self.handle_button_click)
        layout.addWidget(self.button_box)

        self.apply_palette_button.clicked.connect(self._apply_color_palette)

    def _populate_initial_values(self):
        x_lim = self.current_settings["x_lim"] or self.axes.get_xlim();
        y_lim = self.current_settings["y_lim"] or self.axes.get_ylim()
        self.x_label_edit.setText(self.current_settings["x_label"]);
        self.y_label_edit.setText(self.current_settings["y_label"])
        self.x_from.setRange(-10000, 10000);
        self.x_from.setValue(x_lim[0])
        self.x_to.setRange(-10000, 10000);
        self.x_to.setValue(x_lim[1])
        self.y_from.setRange(-10000, 10000);
        self.y_from.setValue(y_lim[0])
        self.y_to.setRange(-10000, 10000);
        self.y_to.setValue(y_lim[1])
        self.x_ticks.setValue(self.current_settings["x_ticks"] or 10.0)
        self.y_ticks.setValue(self.current_settings["y_ticks"] or 0.2)
        self.x_visible.setChecked(self.current_settings["x_visible"]);
        self.y_visible.setChecked(self.current_settings["y_visible"])
        self.grid_visible.setChecked(self.current_settings["grid_visible"])
        self.x_label_size.setValue(self.current_settings["x_label_fontsize"]);
        self.x_label_bold.setChecked(self.current_settings["x_label_bold"]);
        self.x_label_italic.setChecked(self.current_settings["x_label_italic"])
        self.y_label_size.setValue(self.current_settings["y_label_fontsize"]);
        self.y_label_bold.setChecked(self.current_settings["y_label_bold"]);
        self.y_label_italic.setChecked(self.current_settings["y_label_italic"])
        self.global_offset_spin.setValue(self.current_settings.get("global_offset", 0.0))
        self.inter_group_offset_spin.setValue(self.current_settings.get("inter_group_offset", 0.2))
        self.lines_table.setRowCount(len(self.plot_data))
        line_styles = {"Sólida": "-", "Tracejada": "--", "Pontilhada": ":", "Ponto-Traço": "-."}
        for i, data in enumerate(self.plot_data):
            self.lines_table.setItem(i, 0, QTableWidgetItem(data["label"]))
            color_btn = QPushButton();
            color_hex = data.get("color", "#000000")
            color_btn.setStyleSheet(f"background-color: {color_hex};")
            color_btn.clicked.connect(lambda checked=False, btn=color_btn: self.change_line_color(btn))
            self.lines_table.setCellWidget(i, 1, color_btn)
            style_combo = QComboBox();
            style_combo.addItems(line_styles.keys())
            current_style_key = next((k for k, v in line_styles.items() if v == data.get("style", "-")), "Sólida")
            style_combo.setCurrentText(current_style_key)
            self.lines_table.setCellWidget(i, 2, style_combo)
            width_spin = QDoubleSpinBox();
            width_spin.setRange(0.5, 10.0);
            width_spin.setSingleStep(0.1);
            width_spin.setDecimals(2)
            width_spin.setValue(data.get("linewidth", 1.0))
            self.lines_table.setCellWidget(i, 3, width_spin)
            visible_check = QCheckBox();
            visible_check.setChecked(data.get("visible", True))
            visible_check_widget = QWidget()
            chk_layout = QHBoxLayout(visible_check_widget)
            chk_layout.addWidget(visible_check)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            self.lines_table.setCellWidget(i, 4, visible_check_widget)

        self.axes_linewidth_spin.setValue(self.current_settings.get('axes_linewidth', 1.0))

        tick_map = {'in': "Para Dentro", 'out': "Para Fora"}
        self.xtick_direction_combo.setCurrentText(
            tick_map.get(self.current_settings.get('xtick_direction', 'in'), "Sem Ticks"))
        self.ytick_direction_combo.setCurrentText(
            tick_map.get(self.current_settings.get('ytick_direction', 'in'), "Sem Ticks"))

        self.xtick_labelsize_spin.setValue(self.current_settings.get('xtick_labelsize', 10))
        self.ytick_labelsize_spin.setValue(self.current_settings.get('ytick_labelsize', 10))

        self.xtick_visible_check.setChecked(self.current_settings.get('xtick_visible', True))
        self.ytick_visible_check.setChecked(self.current_settings.get('ytick_visible', True))
        self.xtick_label_visible_check.setChecked(self.current_settings.get('xtick_label_visible', True))
        self.ytick_label_visible_check.setChecked(self.current_settings.get('ytick_label_visible', True))
        self.xtick_width_spin.setValue(self.current_settings.get('xtick_width', 0.8))
        self.ytick_width_spin.setValue(self.current_settings.get('ytick_width', 0.8))

    def apply_settings(self):
        # Validar x_ticks e y_ticks para prevenir valores muito pequenos
        x_ticks_value = self.x_ticks.value()
        y_ticks_value = self.y_ticks.value()

        # Se x_ticks ou y_ticks forem muito pequenos (< 0.1), definir como None
        # para usar o locator automático com limite de ticks
        if x_ticks_value > 0 and x_ticks_value < 0.1:
            x_ticks_value = None
            QMessageBox.warning(
                self,
                ptr("Valor Inválido"),
                "O valor de X Ticks é muito pequeno (< 0.1) e foi definido como automático.\n"
                "Valores muito pequenos causam geração excessiva de marcações no eixo."
            )

        if y_ticks_value > 0 and y_ticks_value < 0.1:
            y_ticks_value = None
            QMessageBox.warning(
                self,
                ptr("Valor Inválido"),
                "O valor de Y Ticks é muito pequeno (< 0.1) e foi definido como automático.\n"
                "Valores muito pequenos causam geração excessiva de marcações no eixo."
            )

        axis_settings = {
            "x_lim": (self.x_from.value(), self.x_to.value()), "y_lim": (self.y_from.value(), self.y_to.value()),
            "x_ticks": x_ticks_value if x_ticks_value > 0 else None,
            "y_ticks": y_ticks_value if y_ticks_value > 0 else None,
            "x_visible": self.x_visible.isChecked(), "y_visible": self.y_visible.isChecked(),
            "grid_visible": self.grid_visible.isChecked(),
            "x_label": self.x_label_edit.text(), "y_label": self.y_label_edit.text(),
            "x_label_fontsize": self.x_label_size.value(), "y_label_fontsize": self.y_label_size.value(),
            "x_label_bold": self.x_label_bold.isChecked(), "y_label_bold": self.y_label_bold.isChecked(),
            "x_label_italic": self.x_label_italic.isChecked(), "y_label_italic": self.y_label_italic.isChecked(),
            "global_offset": self.global_offset_spin.value(),
            "inter_group_offset": self.inter_group_offset_spin.value()
        }

        axis_settings['axes_linewidth'] = self.axes_linewidth_spin.value()

        tick_map_inv = {"Para Dentro": 'in', "Para Fora": 'out', "Sem Ticks": None}
        axis_settings['xtick_direction'] = tick_map_inv.get(self.xtick_direction_combo.currentText())
        axis_settings['ytick_direction'] = tick_map_inv.get(self.ytick_direction_combo.currentText())

        axis_settings['xtick_labelsize'] = self.xtick_labelsize_spin.value()
        axis_settings['ytick_labelsize'] = self.ytick_labelsize_spin.value()

        axis_settings['xtick_visible'] = self.xtick_visible_check.isChecked()
        axis_settings['ytick_visible'] = self.ytick_visible_check.isChecked()
        axis_settings['xtick_label_visible'] = self.xtick_label_visible_check.isChecked()
        axis_settings['ytick_label_visible'] = self.ytick_label_visible_check.isChecked()
        axis_settings['xtick_width'] = self.xtick_width_spin.value()
        axis_settings['ytick_width'] = self.ytick_width_spin.value()

        group_spacing = {str(i): spinbox.value() for i, spinbox in self.intra_group_spinboxes.items()}

        line_styles = {"Sólida": "-", "Tracejada": "--", "Pontilhada": ":", "Ponto-Traço": "-."}
        line_settings = []
        for i in range(self.lines_table.rowCount()):
            original_data = self.plot_data[i]
            color_btn = self.lines_table.cellWidget(i, 1)
            color_hex = color_btn.styleSheet().split(":")[1].strip().replace(";", "")
            visible_check_widget = self.lines_table.cellWidget(i, 4)
            visible_check = visible_check_widget.findChild(QCheckBox)
            line_settings.append({
                "id": original_data["id"],
                "label": self.lines_table.item(i, 0).text(),
                "color": color_hex,
                "style": line_styles.get(self.lines_table.cellWidget(i, 2).currentText()),
                "linewidth": self.lines_table.cellWidget(i, 3).value(),
                "visible": visible_check.isChecked()
            })
        self.settings_applied.emit(axis_settings, line_settings, group_spacing)

    def change_line_color(self, button):
        current_color = QColor(button.styleSheet().split(":")[1].strip().replace(";", ""))
        color = QColorDialog.getColor(current_color)
        if color.isValid(): button.setStyleSheet(f"background-color: {color.name()};")

    def handle_button_click(self, button):
        role = self.button_box.buttonRole(button)
        if role == QDialogButtonBox.ButtonRole.ApplyRole:
            self.apply_settings()
        elif role == QDialogButtonBox.ButtonRole.AcceptRole:
            self.apply_settings();
            self.accept()
        elif role == QDialogButtonBox.ButtonRole.RejectRole:
            self.reject()

    def _apply_color_palette(self):
        palette_name_full = self.palette_combo.currentText()
        palette_name = palette_name_full.split(" ")[0].lower().replace("padrão", "default")
        num_items = len(self.plot_data)

        if num_items == 0:
            return

        colors = []
        if palette_name == "default":
            colors = [DEFAULT_PLOT_COLORS[i % len(DEFAULT_PLOT_COLORS)] for i in range(num_items)]
        else:
            try:
                colormap = cm.get_cmap(palette_name, num_items)
                colors = ['#%02x%02x%02x' % (int(r * 255), int(g * 255), int(b * 255)) for r, g, b, a in
                          colormap.colors]
            except Exception as e:
                QMessageBox.critical(self, ptr("Erro de Paleta"),
                                     f"Não foi possível carregar a paleta '{palette_name}':\n{e}")
                return

        for i, color_hex in enumerate(colors):
            color_btn = self.lines_table.cellWidget(i, 1)
            if color_btn:
                color_btn.setStyleSheet(f"background-color: {color_hex};")


class SavePlotDialog(QDialog):
    def __init__(self, default_path="", parent=None):
        super().__init__(parent)
        self.setWindowTitle(ptr("Salvar gráfico"))
        set_polvo_icon(self)
        self.default_path = default_path
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.width_spin = QSpinBox();
        self.width_spin.setRange(100, 8000);
        self.width_spin.setValue(1200);
        self.width_spin.setSuffix(ptr(" px"))
        form_layout.addRow("Largura:", self.width_spin)
        self.height_spin = QSpinBox();
        self.height_spin.setRange(100, 8000);
        self.height_spin.setValue(800);
        self.height_spin.setSuffix(ptr(" px"))
        form_layout.addRow("Altura:", self.height_spin)
        self.dpi_spin = QSpinBox();
        self.dpi_spin.setRange(50, 600);
        self.dpi_spin.setValue(600);
        self.dpi_spin.setSuffix(ptr(" DPI"))
        form_layout.addRow("Qualidade (DPI):", self.dpi_spin)
        self.format_combo = QComboBox();
        self.format_combo.addItems(["png", "jpg", "svg", "pdf", "tiff"])
        form_layout.addRow("Formato:", self.format_combo)
        self.path_edit = QLineEdit();
        self.path_edit.setReadOnly(True)
        browse_button = QPushButton(ptr("Procurar..."));
        browse_button.clicked.connect(self.browse_path)
        path_layout = QHBoxLayout();
        path_layout.addWidget(self.path_edit);
        path_layout.addWidget(browse_button)
        form_layout.addRow("Salvar em:", path_layout)
        layout.addLayout(form_layout)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept);
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def browse_path(self):
        file_format = self.format_combo.currentText()
        file_filter = f"{file_format.upper()} Arquivos (*.{file_format});;Todos os Arquivos (*)"
        path, _ = QFileDialog.getSaveFileName(self, "Salvar Gráfico", self.default_path, file_filter)
        if path: self.path_edit.setText(path)

    def get_settings(self):
        return {"path": self.path_edit.text(), "format": self.format_combo.currentText(),
                "width": self.width_spin.value(), "height": self.height_spin.value(), "dpi": self.dpi_spin.value()}

    def accept(self):
        if not self.path_edit.text():
            QMessageBox.warning(self, ptr("Caminho Inválido"), ptr("Por favor, escolha um local para salvar o arquivo."))
            return
        super().accept()


class SmoothDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(ptr("Parâmetros de Suavização"))
        set_polvo_icon(self)
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.window_spin = QSpinBox();
        self.window_spin.setRange(3, 999);
        self.window_spin.setValue(15);
        self.window_spin.setSingleStep(2)
        form_layout.addRow("Tamanho da Janela (ímpar):", self.window_spin)
        self.order_spin = QSpinBox();
        self.order_spin.setRange(1, 10);
        self.order_spin.setValue(2)
        form_layout.addRow("Ordem do Polinômio:", self.order_spin)
        layout.addLayout(form_layout)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept);
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_settings(self): return {"window": self.window_spin.value(), "order": self.order_spin.value()}

    def accept(self):
        if self.order_spin.value() >= self.window_spin.value():
            QMessageBox.warning(self, ptr("Parâmetros Inválidos"),
                                ptr("A 'Ordem do Polinômio' deve ser menor que o 'Tamanho da Janela'."))
            return
        super().accept()


class PeakDetectDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(ptr("Parâmetros de Detecção de Picos"))
        set_polvo_icon(self)
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.prominence_spin = QDoubleSpinBox();
        self.prominence_spin.setRange(0.001, 1.0);
        self.prominence_spin.setValue(0.02);
        self.prominence_spin.setSingleStep(0.01);
        self.prominence_spin.setDecimals(3)
        form_layout.addRow("Proeminência (0-1):", self.prominence_spin)
        self.width_spin = QSpinBox();
        self.width_spin.setRange(1, 100);
        self.width_spin.setValue(3)
        form_layout.addRow("Largura Mínima (pontos):", self.width_spin)
        layout.addLayout(form_layout)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept);
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_settings(self): return {"prominence": self.prominence_spin.value(), "width": self.width_spin.value()}


# --- NOVA CLASSE DE DIÁLOGO PARA WAVELET DENOISING ---
class WaveletDenoiseDialog(QDialog):
    denoising_applied = Signal(np.ndarray)

    def __init__(self, x_data, y_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle(ptr("Redução de Ruído com Transformada Wavelet"))
        self.setMinimumSize(800, 600)
        set_polvo_icon(self)

        self.x_data = x_data
        self.y_data = y_data
        self.current_denoised_y = self.y_data.copy()

        self._init_ui()
        self._setup_connections()
        self._update_plot()

    def _init_ui(self):
        main_layout = QVBoxLayout(self)
        self.plot_canvas = FigureCanvas(Figure(figsize=(8, 5)))
        self.axes = self.plot_canvas.figure.add_subplot(111)
        main_layout.addWidget(self.plot_canvas)

        controls_group = QGroupBox(ptr("Controles"))
        controls_layout = QFormLayout(controls_group)

        self.wavelet_combo = QComboBox()
        # Lista de algumas wavelets comuns e eficazes
        self.wavelet_combo.addItems(['sym8', 'db4', 'coif5', 'bior3.5'])
        controls_layout.addRow("Tipo de Wavelet:", self.wavelet_combo)

        self.threshold_slider = QSlider(Qt.Orientation.Horizontal)
        self.threshold_slider.setRange(1, 100)  # Mapeado para 0.01 a 1.0
        self.threshold_slider.setValue(10)
        self.threshold_label = QLabel()
        slider_layout = QHBoxLayout()
        slider_layout.addWidget(self.threshold_slider)
        slider_layout.addWidget(self.threshold_label)
        controls_layout.addRow("Nível de Limiar (Threshold):", slider_layout)

        main_layout.addWidget(controls_group)

        self.button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        main_layout.addWidget(self.button_box)

    def _setup_connections(self):
        self.wavelet_combo.currentTextChanged.connect(self._update_plot)
        self.threshold_slider.valueChanged.connect(self._update_plot)
        self.button_box.accepted.connect(self._apply_and_accept)
        self.button_box.rejected.connect(self.reject)

    @Slot()
    def _update_plot(self):
        wavelet = self.wavelet_combo.currentText()
        threshold_value = self.threshold_slider.value() / 100.0
        self.threshold_label.setText(f"{threshold_value:.2f}")

        # --- Alteração: Acesso à função de 'xrd_math_tools' ---
        if not PYWAVELETS_AVAILABLE:
            self.axes.clear()
            self.axes.text(0.5, 0.5, "Erro: Biblioteca PyWavelets não instalada.",
                           horizontalalignment='center', color='red')
            self.plot_canvas.draw()
            return

        self.current_denoised_y = denoise_with_wavelets(self.y_data, wavelet, threshold_value)
        # --- Fim da Alteração ---

        self.axes.clear()
        self.axes.plot(self.x_data, self.y_data, label="Sinal Original", color='blue', alpha=0.5)
        self.axes.plot(self.x_data, self.current_denoised_y, label="Sinal Tratado (Wavelet)", color='red',
                       linewidth=1.5)
        self.axes.set_xlabel("2θ (Graus)")
        self.axes.set_ylabel("Intensidade")
        self.axes.legend()
        self.axes.grid(True, linestyle=':', alpha=0.6)
        self.plot_canvas.figure.tight_layout()
        self.plot_canvas.draw()

    def _apply_and_accept(self):
        self.denoising_applied.emit(self.current_denoised_y)
        self.accept()